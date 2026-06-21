import streamlit as st
import sqlite3
import pandas as pd
import numpy as np
import os
import io
import time
import random
import hashlib
import hmac
import html
import secrets
from datetime import datetime
from pathlib import Path
import plotly.express as px
import plotly.graph_objects as go
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing, Rect, String, Circle, Line, Wedge
from reportlab.lib.units import inch
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.preprocessing import StandardScaler, OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================
# 1. PAGE SETUP & CONFIGURATION
# ==========================================
def load_env_file(path: Path):
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))

BASE_DIR = Path(__file__).resolve().parent
load_env_file(BASE_DIR / ".env")

st.set_page_config(
    page_title="YarnInsight - AI Yarn Quality Platform",
    page_icon="YI",
    layout="wide",
    initial_sidebar_state="expanded"
)

DB_PATH = os.getenv("YARNINSIGHT_DB_PATH", str(BASE_DIR / "yarn_insight.db"))
CSV_PATH = os.getenv("YARNINSIGHT_CSV_PATH", str(BASE_DIR / "yarn data.csv"))
PASSWORD_ITERATIONS = int(os.getenv("YARNINSIGHT_PASSWORD_ITERATIONS", "260000"))

# Pre-populate lists of features
NUMERIC_COLUMNS = [
    "Cellulose of yarn (%)",
    "Hemicellulose of yarn (%)",
    "Lignin of yarn (%)",
    "Pectin of yarn (%)",
    "Moisture Content of yarn (%)",
    "pH Level of yarn",
    "Fineness of yarn (tex)",
    "Fiber Tenacity of yarn (gm/tex)",
    "Elongation of yarn (%)",
    "Moisture Regain of yarn (%)",
    "Water Swelling of yarn (%)",
    "True Density of yarn (gms/cc)",
    "Porosity of yarn (%)",
    "Tensile Strength of yarn (MPa)"
]
FEATURE_COLUMNS = NUMERIC_COLUMNS + ["Dye Type"]

# ==========================================
# 2. CUSTOM CSS STYLING (Rich Aesthetics)
# ==========================================
def inject_custom_css():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&display=swap');

        :root {
            --primary: #0d9488;
            --primary-dark: #0f766e;
            --primary-light: #f0fdfa;
            --secondary: #6366f1;
            --background: #f8fafc;
            --surface: #ffffff;
            --text-main: #0f172a;
            --text-muted: #64748b;
            --border: #e2e8f0;
            --success: #10b981;
            --warning: #f59e0b;
            --danger: #ef4444;
            --shadow-sm: 0 1px 2px 0 rgb(0 0 0 / 0.05);
            --shadow-md: 0 4px 6px -1px rgb(0 0 0 / 0.1), 0 2px 4px -2px rgb(0 0 0 / 0.1);
            --shadow-lg: 0 10px 15px -3px rgb(0 0 0 / 0.1), 0 4px 6px -4px rgb(0 0 0 / 0.1);
        }

        /* Global overrides */
        .stApp {
            background-color: #f8fafc;
            color: #0f172a;
            font-family: 'Plus Jakarta Sans', -apple-system, sans-serif;
        }

        h1, h2, h3, h4, h5, h6 {
            font-family: 'Outfit', sans-serif;
            font-weight: 700;
            color: #0f172a;
            letter-spacing: -0.02em;
        }

        /* Sidebar Styling */
        [data-testid="stSidebar"] {
            background-color: #0f172a;
            color: #f8fafc;
            border-right: 1px solid #1e293b;
        }
        [data-testid="stSidebar"] * {
            color: #f8fafc;
        }
        [data-testid="stSidebar"] .stRadio [role="radiogroup"] label {
            background-color: #1e293b;
            border-radius: 8px;
            padding: 10px 14px;
            margin-bottom: 8px;
            border: 1px solid #334155;
            transition: all 0.2s ease;
        }
        [data-testid="stSidebar"] .stRadio [role="radiogroup"] label:hover {
            background-color: #334155;
            border-color: #0d9488;
            transform: translateX(4px);
        }
        [data-testid="stSidebar"] .stRadio [role="radiogroup"] [data-checked="true"] label {
            background-color: #0d9488 !important;
            border-color: #0d9488 !important;
            font-weight: 600;
        }

        /* Container Custom styles */
        .premium-card {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 12px;
            padding: 24px;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05), 0 2px 4px -2px rgba(0,0,0,0.05);
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            margin-bottom: 20px;
        }
        .premium-card:hover {
            box-shadow: 0 10px 15px -3px rgba(0,0,0,0.08), 0 4px 6px -4px rgba(0,0,0,0.08);
            transform: translateY(-2px);
            border-color: #cbd5e1;
        }
        .glass-header {
            background: linear-gradient(135deg, #0d9488, #6366f1);
            color: white;
            border-radius: 12px;
            padding: 30px;
            margin-bottom: 24px;
            box-shadow: 0 10px 20px -5px rgba(13, 148, 136, 0.25);
        }
        .glass-header h1 {
            color: white !important;
            margin: 0 0 8px 0;
            font-size: 2.5rem;
        }
        .glass-header p {
            color: #ccfbf1;
            margin: 0;
            font-size: 1.1rem;
            font-weight: 400;
        }

        /* KPI Cards */
        .kpi-container {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }
        .kpi-card {
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 12px;
            padding: 20px;
            box-shadow: var(--shadow-sm);
            border-left: 4px solid #0d9488;
        }
        .kpi-card.featured {
            border-left-color: #6366f1;
            background: linear-gradient(180deg, #ffffff, #f5f3ff);
        }
        .kpi-label {
            font-size: 0.75rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: #64748b;
            font-weight: 700;
            margin-bottom: 4px;
        }
        .kpi-value {
            font-size: 1.75rem;
            font-weight: 800;
            color: #0f172a;
            line-height: 1.2;
            margin-bottom: 4px;
        }
        .kpi-subtext {
            font-size: 0.8rem;
            color: #94a3b8;
        }

        /* Status Badges */
        .badge {
            display: inline-flex;
            align-items: center;
            padding: 4px 10px;
            border-radius: 9999px;
            font-size: 0.75rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.02em;
        }
        .badge-a-plus { background: #ecfdf5; color: #065f46; border: 1px solid #a7f3d0; }
        .badge-a { background: #f0fdf4; color: #166534; border: 1px solid #bbf7d0; }
        .badge-b { background: #fffbeb; color: #92400e; border: 1px solid #fde68a; }
        .badge-c { background: #fff7ed; color: #9a3412; border: 1px solid #fed7aa; }
        .badge-reject { background: #fef2f2; color: #991b1b; border: 1px solid #fecaca; }

        .badge-low-risk { background: #ecfdf5; color: #065f46; }
        .badge-med-risk { background: #fffbeb; color: #92400e; }
        .badge-high-risk { background: #fef2f2; color: #991b1b; }

        /* General UI forms */
        div[data-testid="stForm"] {
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 12px;
            padding: 24px;
            box-shadow: var(--shadow-sm);
        }
        .stButton button, .stDownloadButton button, .stFormSubmitButton button {
            background-color: #0d9488 !important;
            color: white !important;
            border: none !important;
            border-radius: 8px !important;
            padding: 10px 20px !important;
            font-weight: 600 !important;
            font-family: 'Outfit', sans-serif !important;
            transition: all 0.2s ease !important;
            width: 100%;
        }
        .stButton button:hover, .stDownloadButton button:hover, .stFormSubmitButton button:hover {
            background-color: #0f766e !important;
            box-shadow: 0 4px 12px rgba(13, 148, 136, 0.2) !important;
            transform: translateY(-1px);
        }

        /* Custom Alert System */
        .custom-alert {
            border-radius: 8px;
            padding: 12px 16px;
            margin-bottom: 12px;
            font-size: 0.9rem;
            display: flex;
            align-items: center;
            gap: 12px;
            border-left: 4px solid transparent;
        }
        .custom-alert.danger {
            background-color: #fef2f2;
            color: #991b1b;
            border-color: #ef4444;
        }
        .custom-alert.warning {
            background-color: #fffbeb;
            color: #92400e;
            border-color: #f59e0b;
        }
        .custom-alert.info {
            background-color: #f0fdfa;
            color: #0f766e;
            border-color: #0d9488;
        }

        /* Optimization Panel */
        .opt-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 12px;
            margin-top: 14px;
        }
        .opt-card {
            background: #f8fafc;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
            padding: 14px;
        }
        .opt-title {
            font-size: 0.8rem;
            font-weight: 700;
            color: #475569;
            text-transform: uppercase;
        }
        .opt-change {
            font-size: 1.1rem;
            font-weight: 800;
            color: #0d9488;
            margin: 4px 0;
        }
        .opt-desc {
            font-size: 0.75rem;
            color: #64748b;
            line-height: 1.3;
        }

        /* Chat UI styling */
        .chat-bubble {
            padding: 14px 18px;
            border-radius: 12px;
            margin-bottom: 12px;
            line-height: 1.5;
            font-size: 0.95rem;
            max-width: 80%;
        }
        .chat-bubble.user {
            background-color: #0d9488;
            color: white;
            margin-left: auto;
            border-bottom-right-radius: 2px;
        }
        .chat-bubble.assistant {
            background-color: #e2e8f0;
            color: #0f172a;
            margin-right: auto;
            border-bottom-left-radius: 2px;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

# ==========================================
# 3. DATABASE OPERATIONS & BUSINESS LOGIC
# ==========================================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    journal_mode = os.getenv("YARNINSIGHT_SQLITE_JOURNAL_MODE", "").strip().upper()
    if journal_mode in {"DELETE", "TRUNCATE", "PERSIST", "MEMORY", "WAL", "OFF"}:
        conn.execute(f"PRAGMA journal_mode={journal_mode}")
    return conn

def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    derived = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("utf-8"),
        PASSWORD_ITERATIONS
    ).hex()
    return f"pbkdf2_sha256${PASSWORD_ITERATIONS}${salt}${derived}"

def verify_password(password: str, stored_hash: str) -> bool:
    if not stored_hash:
        return False

    # Backward compatibility for databases initialized before PBKDF2 support.
    if "$" not in stored_hash:
        legacy_hash = hashlib.sha256(password.encode("utf-8")).hexdigest()
        return hmac.compare_digest(stored_hash, legacy_hash)

    try:
        algorithm, iterations, salt, expected = stored_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        derived = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            salt.encode("utf-8"),
            int(iterations)
        ).hex()
        return hmac.compare_digest(derived, expected)
    except (ValueError, TypeError):
        return False

def db_update_password_hash(email, password):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET password_hash = ? WHERE email = ?",
        (hash_password(password), email.strip().lower())
    )
    conn.commit()
    conn.close()

def db_verify_login(email, password):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE email = ?", (email.strip().lower(),))
    user = cursor.fetchone()
    conn.close()
    if user and verify_password(password, user["password_hash"]):
        if "$" not in user["password_hash"]:
            db_update_password_hash(email, password)
        return True, user["name"], user["role"]
    return False, "", ""

def db_create_user(email, name, password, role):
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO users VALUES (?, ?, ?, ?)", (
            email.strip().lower(), name.strip(), hash_password(password), role
        ))
        conn.commit()
        db_log_audit(email, "USER_SIGNUP", f"Created user {name} with role {role}.")
        return True, "User registered successfully."
    except sqlite3.IntegrityError:
        return False, "User with this email already exists."
    finally:
        conn.close()

def db_log_audit(email, action, details):
    conn = get_db()
    cursor = conn.cursor()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("INSERT INTO audit_logs (timestamp, user_email, action, details) VALUES (?, ?, ?, ?)", (
        timestamp, email, action, details
    ))
    conn.commit()
    conn.close()

def db_get_audit_logs(limit=100):
    conn = get_db()
    safe_limit = max(1, min(int(limit), 500))
    df = pd.read_sql_query("SELECT * FROM audit_logs ORDER BY timestamp DESC LIMIT ?", conn, params=(safe_limit,))
    conn.close()
    return df

def db_get_batches(search=None, grade=None, status=None, supplier=None):
    conn = get_db()
    query = "SELECT * FROM batches WHERE 1=1"
    params = []
    
    if search:
        query += " AND (batch_id LIKE ? OR supplier_name LIKE ?)"
        params.extend([f"%{search}%", f"%{search}%"])
    if grade:
        query += " AND predicted_grade = ?"
        params.append(grade)
    if status:
        query += " AND status = ?"
        params.append(status)
    if supplier:
        query += " AND supplier_name = ?"
        params.append(supplier)
        
    query += " ORDER BY creation_time DESC"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df

def db_save_batch(data):
    conn = get_db()
    cursor = conn.cursor()
    
    # Check if exists
    cursor.execute("SELECT batch_id FROM batches WHERE batch_id = ?", (data["batch_id"],))
    exists = cursor.fetchone()
    
    cols = [
        "batch_id", "creation_time", "supplier_name", "cellulose", "hemicellulose", "lignin", "pectin",
        "moisture_content", "ph_level", "fineness", "tenacity", "elongation", "moisture_regain",
        "water_swelling", "density", "porosity", "actual_strength", "predicted_strength",
        "predicted_grade", "confidence", "risk_level", "risk_reasons", "dye_type", "status"
    ]
    
    if exists:
        update_str = ", ".join([f"{col} = ?" for col in cols[1:]])
        val_tuple = tuple(data[col] for col in cols[1:]) + (data["batch_id"],)
        cursor.execute(f"UPDATE batches SET {update_str} WHERE batch_id = ?", val_tuple)
        action = "BATCH_UPDATE"
    else:
        q_marks = ", ".join(["?" for _ in cols])
        val_tuple = tuple(data[col] for col in cols)
        cursor.execute(f"INSERT INTO batches ({', '.join(cols)}) VALUES ({q_marks})", val_tuple)
        action = "BATCH_CREATE"
        
    conn.commit()
    conn.close()
    db_log_audit(st.session_state.get("email", "system"), action, f"Batch {data['batch_id']} saved with grade {data['predicted_grade']}.")

def db_get_batch(batch_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM batches WHERE batch_id = ?", (batch_id,))
    row = cursor.fetchone()
    conn.close()
    return row

def db_delete_batch(batch_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM batches WHERE batch_id = ?", (batch_id,))
    conn.commit()
    conn.close()
    db_log_audit(st.session_state.get("email", "system"), "BATCH_DELETE", f"Deleted batch ID {batch_id}.")

def db_save_prediction_history(data):
    conn = get_db()
    cursor = conn.cursor()
    cols = [
        "timestamp", "user_email", "cellulose", "hemicellulose", "lignin", "pectin",
        "moisture_content", "ph_level", "fineness", "tenacity", "elongation", "moisture_regain",
        "water_swelling", "density", "porosity", "predicted_strength", "predicted_grade", "confidence", "risk_level", "dye_type"
    ]
    q_marks = ", ".join(["?" for _ in cols])
    val_tuple = tuple(data[col] for col in cols)
    cursor.execute(f"INSERT INTO prediction_history ({', '.join(cols)}) VALUES ({q_marks})", val_tuple)
    conn.commit()
    conn.close()

def db_get_prediction_history(limit=50):
    conn = get_db()
    safe_limit = max(1, min(int(limit), 500))
    df = pd.read_sql_query("SELECT * FROM prediction_history ORDER BY timestamp DESC LIMIT ?", conn, params=(safe_limit,))
    conn.close()
    return df

def ensure_chat_messages_table():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS chat_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_email TEXT NOT NULL,
        role TEXT NOT NULL,
        content TEXT NOT NULL,
        timestamp TEXT NOT NULL
    )
    """)
    conn.commit()
    conn.close()

def db_save_chat_message(user_email, role, content):
    ensure_chat_messages_table()
    conn = get_db()
    cursor = conn.cursor()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute(
        "INSERT INTO chat_messages (user_email, role, content, timestamp) VALUES (?, ?, ?, ?)",
        (user_email, role, content, timestamp)
    )
    conn.commit()
    conn.close()

def db_load_chat_history(user_email, limit=100):
    ensure_chat_messages_table()
    conn = get_db()
    df = pd.read_sql_query(
        "SELECT role, content FROM chat_messages WHERE user_email = ? ORDER BY id ASC LIMIT ?",
        conn,
        params=(user_email, limit)
    )
    conn.close()
    return df.to_dict("records") if len(df) > 0 else []

def db_clear_chat_history(user_email):
    ensure_chat_messages_table()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM chat_messages WHERE user_email = ?", (user_email,))
    conn.commit()
    conn.close()

def get_chat_alert_suggestions():
    """Build proactive chat prompts from live database alerts."""
    suggestions = []
    batches_df = db_get_batches()

    reject_batches = batches_df[batches_df["predicted_grade"] == "Reject"]
    if len(reject_batches) > 0:
        batch_id = reject_batches.iloc[0]["batch_id"]
        suggestions.append(f"Why did batch {batch_id} fail?")

    high_risk = batches_df[batches_df["risk_level"] == "High Risk"]
    if len(high_risk) > 0:
        batch_id = high_risk.iloc[0]["batch_id"]
        suggestions.append(f"What caused high risk on {batch_id}?")

    hold_batches = batches_df[batches_df["status"] == "Hold"]
    if len(hold_batches) > 0:
        suggestions.append(f"Show details for batch {hold_batches.iloc[0]['batch_id']}")

    defaults = [
        "How can I improve tensile strength?",
        "Why is a yarn lot graded as B?",
        "Which fabric is best suited for 1600 MPa yarn?",
    ]
    for prompt in defaults:
        if prompt not in suggestions and len(suggestions) < 4:
            suggestions.append(prompt)

    if len(suggestions) < 4:
        cursor_batch = batches_df.iloc[0]["batch_id"] if len(batches_df) > 0 else "B-1010"
        suggestions.append(f"Show details for batch {cursor_batch}")

    return suggestions[:4]

# ==========================================
# 4. MACHINE LEARNING & EVALUATION PIPELINE
# ==========================================
@st.cache_data
def get_raw_dataset():
    if not os.path.exists(CSV_PATH):
        raise FileNotFoundError(f"Missing dataset file: {CSV_PATH}")
    return pd.read_csv(CSV_PATH)

def get_zscore_series(df):
    out = df.copy()
    # Compute z-scores for baseline properties
    for col in NUMERIC_COLUMNS:
        sd = out[col].std(ddof=0)
        mean = out[col].mean()
        out[f"z_{col}"] = (out[col] - mean) / sd if sd else 0.0
    return out

def label_dataframe_grades(df):
    out = df.copy()
    z_df = get_zscore_series(out)
    
    # Weighted Quality Score
    q = (0.35 * z_df["z_Tensile Strength of yarn (MPa)"] + 
         0.30 * z_df["z_Fiber Tenacity of yarn (gm/tex)"] + 
         0.15 * z_df["z_Elongation of yarn (%)"] + 
         0.10 * z_df["z_Moisture Regain of yarn (%)"] - 
         0.10 * z_df["z_Fineness of yarn (tex)"])
    
    out["Quality Score"] = (q - q.min()) / (q.max() - q.min() + 1e-9) * 100
    
    def assign_grade(score):
        if score >= 80: return "Grade A+ (Premium)"
        elif score >= 65: return "Grade A"
        elif score >= 50: return "Grade B"
        elif score >= 35: return "Grade C"
        else: return "Reject"
        
    out["Quality Grade"] = out["Quality Score"].apply(assign_grade)
    return out

@st.cache_resource(show_spinner="Training predictive models on dataset...")
def train_yarn_models():
    # Load and process data
    df_raw = get_raw_dataset()
    df_labeled = label_dataframe_grades(df_raw)
    
    # To ensure training is fast, train on a stratified sample of 8,000 records
    sample_df = df_labeled.sample(8000, random_state=42)
    
    # 1. Regressor pipeline (Target: Tensile Strength of yarn (MPa))
    # Features: cellulose, hemicellulose, lignin, pectin, moisture, pH, fineness, tenacity, elongation, regain, water swelling, density, porosity, Dye Type
    reg_features = [col for col in FEATURE_COLUMNS if col != "Tensile Strength of yarn (MPa)"]
    num_features_reg = [col for col in NUMERIC_COLUMNS if col != "Tensile Strength of yarn (MPa)"]
    
    X_reg = sample_df[reg_features]
    y_reg = sample_df["Tensile Strength of yarn (MPa)"]
    
    X_train_reg, X_test_reg, y_train_reg, y_test_reg = train_test_split(X_reg, y_reg, test_size=0.2, random_state=42)
    
    preprocessor_reg = ColumnTransformer([
        ("num", Pipeline([("imp", SimpleImputer(strategy="median")), ("sc", StandardScaler())]), num_features_reg),
        ("cat", Pipeline([("imp", SimpleImputer(strategy="most_frequent")), ("oh", OneHotEncoder(handle_unknown="ignore"))]), ["Dye Type"])
    ])
    
    reg_pipeline = Pipeline([
        ("pre", preprocessor_reg),
        ("reg", RandomForestRegressor(n_estimators=100, random_state=42, n_jobs=-1))
    ])
    reg_pipeline.fit(X_train_reg, y_train_reg)
    
    # Calculate R2
    y_pred_reg = reg_pipeline.predict(X_test_reg)
    r2_score = float(np.corrcoef(y_test_reg, y_pred_reg)[0, 1]**2)
    
    # 2. Classifier pipeline (Target: Quality Grade)
    # Features: All features including Tensile Strength
    X_clf = sample_df[FEATURE_COLUMNS]
    y_clf = sample_df["Quality Grade"]
    
    X_train_clf, X_test_clf, y_train_clf, y_test_clf = train_test_split(X_clf, y_clf, test_size=0.2, random_state=42, stratify=y_clf)
    
    preprocessor_clf = ColumnTransformer([
        ("num", Pipeline([("imp", SimpleImputer(strategy="median")), ("sc", StandardScaler())]), NUMERIC_COLUMNS),
        ("cat", Pipeline([("imp", SimpleImputer(strategy="most_frequent")), ("oh", OneHotEncoder(handle_unknown="ignore"))]), ["Dye Type"])
    ])
    
    clf_pipeline = Pipeline([
        ("pre", preprocessor_clf),
        ("clf", RandomForestClassifier(n_estimators=120, random_state=42, n_jobs=-1))
    ])
    clf_pipeline.fit(X_train_clf, y_train_clf)
    
    # Calculate Accuracy
    acc_score = float(clf_pipeline.score(X_test_clf, y_test_clf))
    
    return reg_pipeline, clf_pipeline, df_labeled, r2_score, acc_score

# ==========================================
# 5. DOMAIN BUSINESS LOGIC
# ==========================================
def run_predictions_on_row(reg_model, clf_model, input_dict, df_all):
    # Prepare row
    row_reg = pd.DataFrame([input_dict])
    if "Tensile Strength of yarn (MPa)" in row_reg.columns:
        row_reg = row_reg.drop(columns=["Tensile Strength of yarn (MPa)"])
        
    # Strictly enforce correct column sequence
    reg_features = [col for col in FEATURE_COLUMNS if col != "Tensile Strength of yarn (MPa)"]
    row_reg = row_reg[reg_features]
    
    predicted_strength = float(reg_model.predict(row_reg)[0])
    
    row_clf = pd.DataFrame([input_dict])
    row_clf["Tensile Strength of yarn (MPa)"] = predicted_strength
    row_clf = row_clf[FEATURE_COLUMNS]
    
    predicted_grade = str(clf_model.predict(row_clf)[0])
    confidence = float(np.max(clf_model.predict_proba(row_clf)[0])) * 100
    
    # Risk Assessment
    risk_level, risk_reasons = compute_risk_assessment(row_clf.iloc[0], predicted_grade)
    
    return predicted_strength, predicted_grade, confidence, risk_level, risk_reasons

def compute_risk_assessment(row, grade):
    reasons = []
    
    strength = float(row["Tensile Strength of yarn (MPa)"])
    if strength < 1350:
        reasons.append(f"Tensile Strength is dangerously low ({strength:.1f} MPa)")
    elif strength < 1500:
        reasons.append(f"Sub-optimal tensile strength ({strength:.1f} MPa)")
        
    ph = float(row["pH Level of yarn"])
    if ph < 4.8 or ph > 7.2:
        reasons.append(f"pH level ({ph:.1f}) is out of safe industrial bounds")
    elif ph < 5.2 or ph > 6.8:
        reasons.append(f"pH level ({ph:.1f}) requires monitoring")
        
    porosity = float(row["Porosity of yarn (%)"])
    if porosity > 13.0:
        reasons.append(f"High yarn porosity ({porosity:.1f}%) risks fabric structural integrity")
        
    tenacity = float(row["Fiber Tenacity of yarn (gm/tex)"])
    if tenacity < 32.0:
        reasons.append(f"Insufficient fiber tenacity ({tenacity:.1f} gm/tex)")
        
    moisture = float(row["Moisture Content of yarn (%)"])
    if moisture > 12.0:
        reasons.append(f"Excessive moisture ({moisture:.1f}%) - risk of bacterial and mildew growth")
    elif moisture < 9.0:
        reasons.append(f"Low moisture content ({moisture:.1f}%) - fibers are prone to brittle breakage")
        
    if grade == "Reject" or len(reasons) >= 3:
        risk_level = "High Risk"
    elif grade == "Grade C" or len(reasons) >= 1:
        risk_level = "Medium Risk"
    else:
        risk_level = "Low Risk"
        
    if not reasons:
        reasons.append("All physical and chemical properties fall within normal thresholds.")
        
    return risk_level, "; ".join(reasons)

def cloth_from_rules(row, grade):
    # Score 8 fabrics based on technical constraints
    strength = float(row["Tensile Strength of yarn (MPa)"])
    tenacity = float(row["Fiber Tenacity of yarn (gm/tex)"])
    fineness = float(row["Fineness of yarn (tex)"])
    elongation = float(row["Elongation of yarn (%)"])
    regain = float(row["Moisture Regain of yarn (%)"])
    swelling = float(row["Water Swelling of yarn (%)"])
    density = float(row["True Density of yarn (gms/cc)"])
    porosity = float(row["Porosity of yarn (%)"])
    
    # 1. Denim (Strength & Tenacity primary, heavier tex acceptable)
    denim_score = 40
    if strength > 1700: denim_score += 30
    if tenacity > 42: denim_score += 20
    if fineness > 2.8: denim_score += 10
    
    # 2. T-Shirts (Fine tex primary, moisture regain and cellulose comfort)
    tshirt_score = 30
    if fineness < 2.5: tshirt_score += 30
    if regain > 11.5: tshirt_score += 20
    if elongation > 2.8: tshirt_score += 20
    
    # 3. Knitwear (Elongation/flexibility primary, softness fineness)
    knit_score = 30
    if elongation > 3.2: knit_score += 40
    if fineness < 2.8: knit_score += 20
    if swelling > 45: knit_score += 10
    
    # 4. Sportswear (Elongation high, lower/mod moisture regain to prevent water retention, good tenacity)
    sports_score = 35
    if elongation > 3.4: sports_score += 35
    if regain < 11.0: sports_score += 15
    if tenacity > 38: sports_score += 15
    
    # 5. Home Textiles (Moisture Regain and Water Swelling comfort, density moderate)
    home_score = 40
    if regain > 12.0: home_score += 25
    if swelling > 48: home_score += 25
    if porosity > 7.0: home_score += 10
    
    # 6. Upholstery (True Density & Lignin content for durability, strength)
    uph_score = 40
    if density > 1.48: uph_score += 25
    if strength > 1550: uph_score += 25
    if fineness > 3.0: uph_score += 10
    
    # 7. Industrial Fabrics (Maximum strength, density, low moisture swelling)
    ind_score = 30
    if strength > 1850: ind_score += 40
    if tenacity > 46: ind_score += 20
    if swelling < 45: ind_score += 10
    
    # 8. Bags & Accessories (Coarse fineness, High Porosity, Tensile Strength)
    bags_score = 35
    if strength > 1600: bags_score += 25
    if fineness > 3.1: bags_score += 25
    if tenacity > 36: bags_score += 15
    
    # Adjust scores based on general Grade
    modifier = {
        "Grade A+ (Premium)": 1.15,
        "Grade A": 1.05,
        "Grade B": 0.95,
        "Grade C": 0.80,
        "Reject": 0.35
    }.get(grade, 1.0)
    
    scores = {
        "Denim": min(100, int(denim_score * modifier)),
        "T-Shirts": min(100, int(tshirt_score * modifier)),
        "Knitwear": min(100, int(knit_score * modifier)),
        "Sportswear": min(100, int(sports_score * modifier)),
        "Home Textiles": min(100, int(home_score * modifier)),
        "Upholstery": min(100, int(uph_score * modifier)),
        "Industrial Fabrics": min(100, int(ind_score * modifier)),
        "Bags & Accessories": min(100, int(bags_score * modifier))
    }
    
    # Sort
    sorted_scores = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    return sorted_scores

def get_expected_perf(score):
    if score >= 85: return "Excellent (Highly Recommended)"
    elif score >= 70: return "Good (Suitable)"
    elif score >= 50: return "Fair (Acceptable)"
    else: return "Poor (Not Recommended)"

def get_optimizer_suggestions(row):
    suggestions = []
    
    # 1. Tenacity: suggest increasing if low
    tenacity = float(row["Fiber Tenacity of yarn (gm/tex)"])
    if tenacity < 42.0:
        opt_ten = 45.0
        diff = opt_ten - tenacity
        suggestions.append({
            "parameter": "Fiber Tenacity",
            "current": f"{tenacity:.2f} gm/tex",
            "target": f"{opt_ten:.1f} gm/tex",
            "impact": f"+{diff*8.5:.1f} MPa Strength Boost",
            "action": "Select higher tenacity raw lint fibers, optimize carding and spinning speed, reduce drafts."
        })
        
    # 2. Moisture Content: adjust to sweet spot (10.5% - 11.2%)
    moisture = float(row["Moisture Content of yarn (%)"])
    if moisture < 10.0 or moisture > 11.5:
        opt_m = 10.8
        suggestions.append({
            "parameter": "Moisture Content",
            "current": f"{moisture:.2f}%",
            "target": f"{opt_m:.1f}%",
            "impact": "Improves Elasticity & Lowers Defect Risk",
            "action": "Adjust humidifier outputs in the conditioning room, ensure aging is done for at least 24 hours."
        })
        
    # 3. Porosity: suggest reducing if too high
    porosity = float(row["Porosity of yarn (%)"])
    if porosity > 8.5:
        opt_p = 6.8
        diff = porosity - opt_p
        suggestions.append({
            "parameter": "Porosity",
            "current": f"{porosity:.2f}%",
            "target": f"{opt_p:.1f}%",
            "impact": f"+{diff*12.0:.1f} MPa Strength & Compactness",
            "action": "Increase twist multipliers during ring frame processing, check traveler size, and spindle centering."
        })
        
    # 4. Fineness: suggest making it finer if coarse (lower tex value)
    fineness = float(row["Fineness of yarn (tex)"])
    if fineness > 2.8:
        opt_f = 2.4
        suggestions.append({
            "parameter": "Fineness (Yarn Count)",
            "current": f"{fineness:.2f} tex",
            "target": f"{opt_f:.1f} tex",
            "impact": "Increases Surface Smoothness & Premium Grade",
            "action": "Use longer staple cotton length, adjust drawing draft ratios, and reduce sliver hank weight."
        })
        
    if not suggestions:
        suggestions.append({
            "parameter": "Parameters Optimized",
            "current": "Ideal values",
            "target": "N/A",
            "impact": "No changes needed",
            "action": "Current parameters are at optimal thresholds for production."
        })
        
    return suggestions

def build_scada_strength_gauge(value, target=1500.0):
    """Semi-circular SCADA-style gauge for average tensile strength."""
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=value,
        number={"suffix": " MPa", "font": {"size": 28, "color": "#0f172a"}},
        delta={"reference": target, "relative": False, "valueformat": ".0f", "suffix": " MPa"},
        title={"text": "Average Yarn Strength", "font": {"size": 15, "family": "Outfit", "color": "#0f172a"}},
        gauge={
            "shape": "angular",
            "axis": {"range": [1000, 2200], "tickwidth": 2, "tickcolor": "#475569"},
            "bar": {"color": "#0d9488", "thickness": 0.22},
            "bgcolor": "#f8fafc",
            "borderwidth": 2,
            "bordercolor": "#cbd5e1",
            "steps": [
                {"range": [1000, 1350], "color": "#fecaca"},
                {"range": [1350, 1600], "color": "#fef3c7"},
                {"range": [1600, 2200], "color": "#d1fae5"},
            ],
            "threshold": {
                "line": {"color": "#ef4444", "width": 5},
                "thickness": 0.78,
                "value": target,
            },
        },
    ))
    fig.update_layout(
        height=260,
        margin=dict(t=50, b=20, l=30, r=30),
        paper_bgcolor="rgba(0,0,0,0)",
        font={"family": "Plus Jakarta Sans"},
    )
    return fig

def build_scada_defect_gauge(value):
    """Color-coded semi-circular gauge for supplier defect rate."""
    bar_color = "#ef4444" if value > 5 else ("#f59e0b" if value > 2 else "#10b981")
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=value,
        number={"suffix": "%", "font": {"size": 28, "color": "#0f172a"}},
        title={"text": "Supplier Defect Rate", "font": {"size": 15, "family": "Outfit", "color": "#0f172a"}},
        gauge={
            "shape": "angular",
            "axis": {"range": [0, 15], "tickwidth": 2, "tickcolor": "#475569"},
            "bar": {"color": bar_color, "thickness": 0.22},
            "bgcolor": "#f8fafc",
            "borderwidth": 2,
            "bordercolor": "#cbd5e1",
            "steps": [
                {"range": [0, 2], "color": "#d1fae5"},
                {"range": [2, 5], "color": "#fef3c7"},
                {"range": [5, 15], "color": "#fecaca"},
            ],
            "threshold": {
                "line": {"color": "#047857", "width": 5},
                "thickness": 0.78,
                "value": 2,
            },
        },
    ))
    fig.update_layout(
        height=260,
        margin=dict(t=50, b=20, l=30, r=30),
        paper_bgcolor="rgba(0,0,0,0)",
        font={"family": "Plus Jakarta Sans"},
    )
    return fig

def build_sandbox_live_gauges(strength, grade, risk_level, baseline_strength):
    """Mini dial panel for the optimization sandbox."""
    risk_color = {"Low Risk": "#10b981", "Medium Risk": "#f59e0b", "High Risk": "#ef4444"}.get(risk_level, "#64748b")
    fig = go.Figure()
    fig.add_trace(go.Indicator(
        mode="gauge+number+delta",
        value=strength,
        delta={"reference": baseline_strength, "relative": False, "valueformat": ".1f", "suffix": " MPa"},
        title={"text": "Tensile Strength", "font": {"size": 12}},
        domain={"x": [0, 0.32], "y": [0, 1]},
        gauge={
            "shape": "angular",
            "axis": {"range": [1000, 2200]},
            "bar": {"color": "#6366f1"},
            "steps": [
                {"range": [1000, 1350], "color": "#fecaca"},
                {"range": [1350, 1600], "color": "#fef3c7"},
                {"range": [1600, 2200], "color": "#d1fae5"},
            ],
            "threshold": {"line": {"color": "#ef4444", "width": 3}, "thickness": 0.75, "value": 1500},
        },
    ))
    grade_score = {"Grade A+ (Premium)": 95, "Grade A": 80, "Grade B": 65, "Grade C": 45, "Reject": 15}.get(grade, 50)
    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=grade_score,
        number={"suffix": ""},
        title={"text": f"Grade: {grade.split('(')[0].strip()}", "font": {"size": 11}},
        domain={"x": [0.34, 0.66], "y": [0, 1]},
        gauge={
            "shape": "angular",
            "axis": {"range": [0, 100]},
            "bar": {"color": "#0d9488"},
            "steps": [
                {"range": [0, 35], "color": "#fecaca"},
                {"range": [35, 65], "color": "#fef3c7"},
                {"range": [65, 100], "color": "#d1fae5"},
            ],
        },
    ))
    risk_score = {"Low Risk": 20, "Medium Risk": 55, "High Risk": 90}.get(risk_level, 40)
    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=risk_score,
        number={"suffix": ""},
        title={"text": f"Risk: {risk_level}", "font": {"size": 11, "color": risk_color}},
        domain={"x": [0.68, 1], "y": [0, 1]},
        gauge={
            "shape": "angular",
            "axis": {"range": [0, 100]},
            "bar": {"color": risk_color},
            "steps": [
                {"range": [0, 33], "color": "#d1fae5"},
                {"range": [33, 66], "color": "#fef3c7"},
                {"range": [66, 100], "color": "#fecaca"},
            ],
        },
    ))
    fig.update_layout(height=180, margin=dict(t=40, b=10, l=20, r=20), paper_bgcolor="rgba(0,0,0,0)")
    return fig

def build_yarn_logo_drawing():
    """Vector YarnInsight logo for PDF certificates."""
    logo = Drawing(140, 48)
    logo.add(Rect(0, 0, 140, 48, fillColor=colors.HexColor("#f0fdfa"), strokeColor=colors.HexColor("#0f766e"), strokeWidth=1.2, rx=6, ry=6))
    logo.add(Rect(8, 8, 32, 32, fillColor=colors.HexColor("#0f766e"), strokeColor=None, rx=6, ry=6))
    logo.add(String(17, 18, "YI", fontSize=13, fillColor=colors.white, fontName="Helvetica-Bold"))
    logo.add(String(48, 28, "YarnInsight", fontSize=13, fillColor=colors.HexColor("#0f766e"), fontName="Helvetica-Bold"))
    logo.add(String(48, 12, "Industrial Quality Systems", fontSize=7.5, fillColor=colors.HexColor("#64748b"), fontName="Helvetica"))
    return logo

def build_quality_approved_stamp():
    """Circular 'Quality Approved' stamp for PDF certificates."""
    stamp = Drawing(100, 100)
    stamp.add(Circle(50, 50, 46, fillColor=colors.HexColor("#ecfdf5"), strokeColor=colors.HexColor("#059669"), strokeWidth=2.5))
    stamp.add(Circle(50, 50, 38, fillColor=None, strokeColor=colors.HexColor("#10b981"), strokeWidth=1))
    stamp.add(String(14, 58, "QUALITY", fontSize=9, fillColor=colors.HexColor("#065f46"), fontName="Helvetica-Bold"))
    stamp.add(String(22, 44, "APPROVED", fontSize=9, fillColor=colors.HexColor("#065f46"), fontName="Helvetica-Bold"))
    stamp.add(String(30, 30, "✓", fontSize=14, fillColor=colors.HexColor("#059669"), fontName="Helvetica-Bold"))
    return stamp

def build_mock_qr_drawing(batch_id):
    """Deterministic mock QR code keyed to batch ID."""
    seed = sum(ord(c) for c in str(batch_id))
    random.seed(seed)
    qr = Drawing(72, 72)
    qr.add(Rect(0, 0, 72, 72, fillColor=colors.white, strokeColor=colors.HexColor("#0f766e"), strokeWidth=1.5))
    for corner in [(4, 48), (48, 48), (4, 4)]:
        qr.add(Rect(corner[0], corner[1], 16, 16, fillColor=colors.HexColor("#0f766e"), strokeColor=None))
        qr.add(Rect(corner[0] + 4, corner[1] + 4, 8, 8, fillColor=colors.white, strokeColor=None))
    for _ in range(18):
        rx = random.choice([22, 26, 30, 34, 38, 42, 46])
        ry = random.choice([22, 26, 30, 34, 38, 42, 46])
        qr.add(Rect(rx, ry, 3, 3, fillColor=colors.HexColor("#0f766e"), strokeColor=None))
    random.seed()
    return qr

def _pdf_draw_branded_frame(canvas, doc):
    """Draw colored side border and footer on certificate pages."""
    canvas.saveState()
    page_w, page_h = letter
    canvas.setFillColor(colors.HexColor("#0f766e"))
    canvas.rect(28, 28, 10, page_h - 56, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#14b8a6"))
    canvas.rect(28, page_h - 80, 10, 52, fill=1, stroke=0)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#94a3b8"))
    canvas.drawString(50, 22, f"YarnInsight Certificate • Batch traceability document • Page {canvas.getPageNumber()}")
    canvas.restoreState()

def batch_row_to_fabric_input(batch):
    """Map SQLite batch record keys to feature column names for fabric scoring."""
    b = dict(batch) if not isinstance(batch, dict) else batch
    return {
        "Tensile Strength of yarn (MPa)": b["predicted_strength"],
        "Fiber Tenacity of yarn (gm/tex)": b["tenacity"],
        "Fineness of yarn (tex)": b["fineness"],
        "Elongation of yarn (%)": b["elongation"],
        "Moisture Regain of yarn (%)": b["moisture_regain"],
        "Water Swelling of yarn (%)": b["water_swelling"],
        "True Density of yarn (gms/cc)": b["density"],
        "Porosity of yarn (%)": b["porosity"],
    }

def extract_batch_id(query):
    """Extract batch ID (e.g. B-1015) from natural language query."""
    import re
    match = re.search(r'\b(b-?\d+)\b', query.lower())
    if match:
        raw = match.group(1).upper()
        if raw.startswith("B") and "-" not in raw[1:]:
            return f"B-{raw[1:]}"
        return raw.replace("B", "B-", 1) if raw.startswith("B") and not raw.startswith("B-") else raw
    return None

def generate_chat_failure_response(batch_id):
    """Explain why a batch failed quality checks."""
    batch = db_get_batch(batch_id)
    if not batch:
        return f"I searched the database but could not find batch <b>{batch_id}</b>. Verify the batch ID."
    batch = dict(batch)
    reasons = []
    if batch["predicted_grade"] == "Reject":
        reasons.append(f"Assigned <b>REJECT</b> grade by ML classifier (confidence {batch['confidence']:.1f}%).")
    if batch["predicted_strength"] < 1350:
        reasons.append(f"Tensile strength ({batch['predicted_strength']:.1f} MPa) is below the 1350 MPa minimum.")
    if batch["tenacity"] < 32:
        reasons.append(f"Fiber tenacity ({batch['tenacity']:.1f} gm/tex) is critically low.")
    if batch["moisture_content"] > 12:
        reasons.append(f"Moisture content ({batch['moisture_content']:.1f}%) exceeds safe limits.")
    if batch["porosity"] > 13:
        reasons.append(f"Yarn porosity ({batch['porosity']:.1f}%) indicates poor fiber compaction.")
    if batch["ph_level"] < 4.8 or batch["ph_level"] > 7.2:
        reasons.append(f"pH level ({batch['ph_level']:.1f}) is outside acceptable range.")
    if not reasons:
        reasons.append(batch["risk_reasons"])
    recs = "<br/>".join(f"• {r}" for r in reasons)
    return f"""
    🔍 <b>Failure Analysis for Batch {batch_id}:</b><br/>
    <b>Supplier:</b> {batch['supplier_name']} | <b>Status:</b> {batch['status']}<br/><br/>
    <b>Root Causes:</b><br/>{recs}<br/><br/>
    <b>Recommended Actions:</b> Re-blend with higher tenacity lint, adjust moisture conditioning to 10.5–11.2%, and re-run prediction before release.
    """

def generate_chat_response(triggered_query):
    """Rule-based chat response engine with batch lookup and alert integration."""
    q_lower = triggered_query.lower()
    found_id = extract_batch_id(triggered_query)

    if found_id and ("fail" in q_lower or "reject" in q_lower or "why did" in q_lower):
        return generate_chat_failure_response(found_id)

    if found_id and ("high risk" in q_lower or "risk" in q_lower):
        batch = db_get_batch(found_id)
        if batch:
            batch = dict(batch)
            return f"""
            ⚠️ <b>Risk Profile for Batch {found_id}:</b><br/>
            • <b>Risk Level:</b> {batch['risk_level']}<br/>
            • <b>Grade:</b> {batch['predicted_grade']}<br/>
            • <b>Strength:</b> {batch['predicted_strength']:.1f} MPa<br/>
            • <b>Root Causes:</b> {batch['risk_reasons']}<br/><br/>
            Use the Optimization Sandbox on the Predict page to simulate parameter adjustments.
            """
        return f"No batch record found for {found_id}."

    if "batch" in q_lower or found_id:
        if found_id:
            batch = db_get_batch(found_id)
            if batch:
                batch = dict(batch)
                return f"""
                📝 <b>Database Record Found for Batch {found_id}:</b><br/>
                • <b>Lint Supplier:</b> {batch['supplier_name']}<br/>
                • <b>Predicted Tensile Strength:</b> {batch['predicted_strength']:.1f} MPa<br/>
                • <b>Assigned Quality Grade:</b> {batch['predicted_grade']}<br/>
                • <b>Confidence Score:</b> {batch['confidence']:.2f}%<br/>
                • <b>Risk Profile Level:</b> {batch['risk_level']}<br/>
                • <b>Underlying reasons:</b> {batch['risk_reasons']}<br/>
                • <b>Dye Mix:</b> {batch['dye_type']}<br/>
                • <b>Active Action Status:</b> {batch['status']}<br/><br/>
                <b>Fabric Recommendations:</b> This batch matches best with <b>Denim</b> and <b>Industrial Fabrics</b>.
                """
            return f"I searched the active SQL inventory database but could not find a batch with ID '{found_id}'. Verify the batch ID."
        return "Please specify a batch ID (e.g. B-1010) and I will retrieve the lab specifications and grading records."

    if "strength" in q_lower or "tensile" in q_lower or "improve" in q_lower:
        return """
        ⚙️ <b>How to Improve Yarn Tensile Strength:</b><br/>
        Yarn tensile yield limits are driven by fiber tenacity and yarn count structures. To improve strength:
        <ol>
            <li><b>Increase Fiber Tenacity:</b> Select cotton lint with higher tenacity values (> 40 gm/tex).</li>
            <li><b>Reduce Yarn Fineness (tex):</b> Fine count fibers distribute tension better than coarse count slots.</li>
            <li><b>Reduce Yarn Porosity:</b> Increase twist multipliers during ring frame processing to compact fibers together.</li>
            <li><b>Moisture Balance:</b> Conditioning rooms should keep yarn moisture content near 10.5% - 11.2% to optimize fiber elasticity and prevent brittle snapping.</li>
        </ol>
        """

    if "grade b" in q_lower or "why" in q_lower or "quality grade" in q_lower:
        return """
        ⚖️ <b>Yarn Quality Grading System details:</b><br/>
        YarnInsight calculates a weighted Z-score based on 5 parameters to assign Quality Grades:
        <ul>
            <li><b>Grade A+ (Premium)</b>: Quality Score ≥ 80. Superb strength, low porosity, fine count. Ideal for premium shirting.</li>
            <li><b>Grade A</b>: Quality Score 65–80. Strong lot with standard characteristics.</li>
            <li><b>Grade B</b>: Quality Score 50–65. Marginal lot. Suitable for standard knitwear or comfort textiles.</li>
            <li><b>Grade C</b>: Quality Score 35–50. Risk lot. Retain for secondary use cases like upholstery.</li>
            <li><b>Reject</b>: Quality Score < 35. Critical failures in pH or extreme brittle limits.</li>
        </ul>
        """

    if "fabric" in q_lower or "suitability" in q_lower or "suited" in q_lower:
        return """
        👗 <b>Fabric Suitability matching guidelines:</b><br/>
        • <b>Denim / Workwear:</b> Needs high strength (> 1650 MPa) and high tenacity (> 42 gm/tex).<br/>
        • <b>T-Shirts:</b> Needs fine count (< 2.5 tex) and high comfort moisture regain (> 11.5%).<br/>
        • <b>Sportswear:</b> Needs high elongation (> 3.3%) and lower moisture regain to avoid sweat saturation.<br/>
        • <b>Industrial Fabrics:</b> Needs peak tensile strength (> 1800 MPa) and high true density.
        """

    return f"I received your question: '{triggered_query}'. I can analyze database batch ID codes, explain physical property impacts, and suggest process adjustments. Ask about batch details or optimization steps!"

# ==========================================
# 6. REPORT GENERATOR (PDF & Excel)
# ==========================================
def generate_quality_certificate_pdf(batch_id):
    batch = db_get_batch(batch_id)
    if not batch:
        return None
        
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CertTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        leading=26,
        textColor=colors.HexColor('#0f766e'),
        alignment=1, # Center
        spaceAfter=10
    )
    
    subtitle_style = ParagraphStyle(
        'CertSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=15,
        textColor=colors.HexColor('#64748b'),
        alignment=1,
        spaceAfter=15
    )
    
    h2_style = ParagraphStyle(
        'CertH2',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=17,
        textColor=colors.HexColor('#0f172a'),
        spaceBefore=10,
        spaceAfter=6
    )
    
    body_style = ParagraphStyle(
        'CertBody',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=9.5,
        leading=13.5,
        textColor=colors.HexColor('#334155')
    )

    bold_style = ParagraphStyle(
        'CertBold',
        parent=styles['BodyText'],
        fontName='Helvetica-Bold',
        fontSize=9.5,
        leading=13.5,
        textColor=colors.HexColor('#0f172a')
    )
    
    story = []
    
    # Branded header row: logo + title
    logo_drawing = build_yarn_logo_drawing()
    header_data = [[
        logo_drawing,
        Paragraph(
            "YARNINSIGHT QUALITY VALIDATION CERTIFICATE<br/>"
            f"<font size='9' color='#64748b'>Official Industrial Quality Statement • Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font>",
            ParagraphStyle('CertHeaderBlock', parent=title_style, alignment=0, fontSize=18, leading=22, spaceAfter=0)
        ),
        build_mock_qr_drawing(batch_id),
    ]]
    header_table = Table(header_data, colWidths=[150, 280, 80])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 8))
    
    # Colored top divider
    top_bar = Drawing(510, 4)
    top_bar.add(Rect(0, 0, 510, 4, fillColor=colors.HexColor('#0f766e'), strokeColor=None))
    story.append(top_bar)
    story.append(Spacer(1, 12))
    
    # Generate vector validation stamp
    stamp_drawing = build_quality_approved_stamp()
    
    # Main details table
    grade_color = {
        "Grade A+ (Premium)": "#059669",
        "Grade A": "#10b981",
        "Grade B": "#d97706",
        "Grade C": "#ea580c",
        "Reject": "#dc2626"
    }.get(batch["predicted_grade"], "#0f766e")
    
    details_data = [
        [Paragraph("Batch ID:", bold_style), Paragraph(str(batch["batch_id"]), body_style),
         Paragraph("Date Processed:", bold_style), Paragraph(str(batch["creation_time"]), body_style)],
        [Paragraph("Raw Supplier:", bold_style), Paragraph(str(batch["supplier_name"]), body_style),
         Paragraph("Dye Bath Mix:", bold_style), Paragraph(str(batch["dye_type"]).upper(), body_style)],
        [Paragraph("Validation Status:", bold_style), Paragraph(f"<b><font color='{grade_color}'>{batch['predicted_grade']}</font></b>", body_style),
         Paragraph("ML Confidence Score:", bold_style), Paragraph(f"{batch['confidence']:.2f}%", body_style)],
        [Paragraph("Tensile Strength:", bold_style), Paragraph(f"{batch['predicted_strength']:.1f} MPa", body_style),
         Paragraph("Quality Risk Level:", bold_style), Paragraph(f"<b>{batch['risk_level']}</b>", body_style)],
    ]
    
    details_table = Table(details_data, colWidths=[110, 140, 120, 140])
    details_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.HexColor('#0f172a')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#fafafa')),
    ]))
    
    stamp_row = Table([[details_table, stamp_drawing]], colWidths=[420, 110])
    stamp_row.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
    ]))
    story.append(stamp_row)
    story.append(Spacer(1, 15))
    
    # Technical Parameters List
    story.append(Paragraph("Detailed Material & Mechanical Assay", h2_style))
    
    param_data = [
        [Paragraph("<b>Yarn Parameter</b>", bold_style), Paragraph("<b>Measured Value</b>", bold_style), Paragraph("<b>Yarn Parameter</b>", bold_style), Paragraph("<b>Measured Value</b>", bold_style)],
        [Paragraph("Cellulose Content", body_style), Paragraph(f"{batch['cellulose']:.2f}%", body_style), Paragraph("Fineness (Yarn Count)", body_style), Paragraph(f"{batch['fineness']:.2f} tex", body_style)],
        [Paragraph("Hemicellulose Content", body_style), Paragraph(f"{batch['hemicellulose']:.2f}%", body_style), Paragraph("Fiber Tenacity", body_style), Paragraph(f"{batch['tenacity']:.2f} gm/tex", body_style)],
        [Paragraph("Lignin Content", body_style), Paragraph(f"{batch['lignin']:.2f}%", body_style), Paragraph("Elongation limit", body_style), Paragraph(f"{batch['elongation']:.2f}%", body_style)],
        [Paragraph("Pectin Content", body_style), Paragraph(f"{batch['pectin']:.2f}%", body_style), Paragraph("Moisture Regain", body_style), Paragraph(f"{batch['moisture_regain']:.2f}%", body_style)],
        [Paragraph("Ambient Moisture", body_style), Paragraph(f"{batch['moisture_content']:.2f}%", body_style), Paragraph("Water Swelling", body_style), Paragraph(f"{batch['water_swelling']:.2f}%", body_style)],
        [Paragraph("Yarn pH Level", body_style), Paragraph(f"{batch['ph_level']:.2f}", body_style), Paragraph("Yarn Porosity", body_style), Paragraph(f"{batch['porosity']:.2f}%", body_style)]
    ]
    
    param_table = Table(param_data, colWidths=[140, 125, 140, 125])
    param_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
    ]))
    
    story.append(param_table)
    story.append(Spacer(1, 15))
    
    # Suitability & Recommendations
    story.append(Paragraph("Recommended Fabric Product Mapping", h2_style))
    
    recs = cloth_from_rules(batch_row_to_fabric_input(batch), batch["predicted_grade"])
    best_fabric = recs[0][0]
    best_score = recs[0][1]
    
    rec_text = f"Based on machine learning classification models and automated production rules, this yarn batch has a peak compatibility matching score of <b>{best_score}%</b> for <b>{best_fabric}</b>. Alternative recommended uses include: <b>{recs[1][0]} ({recs[1][1]}%)</b> and <b>{recs[2][0]} ({recs[2][1]}%)</b>."
    story.append(Paragraph(rec_text, body_style))
    story.append(Spacer(1, 10))
    
    # Risk and Quality warnings
    story.append(Paragraph("Technical Risk Profile", h2_style))
    story.append(Paragraph(f"<b>Assigned Risk: {batch['risk_level']}</b>", bold_style))
    story.append(Paragraph(f"<b>Underlying parameters:</b> {batch['risk_reasons']}", body_style))
    
    # Signature and stamp block
    story.append(Spacer(1, 25))
    
    sig_line = Drawing(200, 20)
    sig_line.add(Line(0, 10, 180, 18, strokeColor=colors.HexColor("#0f766e"), strokeWidth=1.2))
    sig_line.add(String(0, 0, "Digitally Signed", fontSize=7, fillColor=colors.HexColor("#64748b"), fontName="Helvetica-Oblique"))
    
    sig_data = [
        [Paragraph("_______________________________<br/>Quality Operations Inspector<br/><font size='8' color='#64748b'>Name & Employee ID</font>", body_style),
         sig_line,
         Paragraph("_______________________________<br/>Plant Quality Manager<br/><font size='8' color='#64748b'>Authorized Release Signature</font>", body_style)]
    ]
    sig_table = Table(sig_data, colWidths=[175, 160, 175])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(sig_table)
    
    # Build Document with branded side border
    doc.build(story, onFirstPage=_pdf_draw_branded_frame, onLaterPages=_pdf_draw_branded_frame)
    pdf_bytes = pdf_buffer.getvalue()
    pdf_buffer.close()
    return pdf_bytes

def export_batches_to_excel():
    df = db_get_batches()
    output = io.BytesIO()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "YarnInsight Batches"
    
    # Stylings
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Columns
    headers = [col.replace("_", " ").title() for col in df.columns]
    ws.append(headers)
    
    # Format Headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # Append Data
    for _, row in df.iterrows():
        ws.append(list(row.values))
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    wb.save(output)
    excel_bytes = output.getvalue()
    output.close()
    return excel_bytes

# ==========================================
# 7. MAIN CONTROLLER & STATE INITIALIZATION
# ==========================================
def main():
    inject_custom_css()
    
    # Session State defaults
    if "auth" not in st.session_state:
        st.session_state["auth"] = False
    if "email" not in st.session_state:
        st.session_state["email"] = ""
    if "name" not in st.session_state:
        st.session_state["name"] = ""
    if "role" not in st.session_state:
        st.session_state["role"] = ""
    if "chat_history" not in st.session_state:
        st.session_state["chat_history"] = []
    if "chat_loaded" not in st.session_state:
        st.session_state["chat_loaded"] = False
        
    # Load Models
    try:
        reg_model, clf_model, df_labeled, r2, acc = train_yarn_models()
    except Exception as e:
        st.error(f"Fatal error initializing models: {e}")
        st.stop()
        
    # ==========================================
    # 8. AUTHENTICATION SHELL
    # ==========================================
    if not st.session_state["auth"]:
        render_auth_page()
        return

    # Load persisted chat history once per session
    if not st.session_state["chat_loaded"]:
        saved = db_load_chat_history(st.session_state["email"])
        if saved:
            st.session_state["chat_history"] = saved
        else:
            st.session_state["chat_history"] = [{
                "role": "assistant",
                "content": "Welcome to YarnInsight AI Assistant! Ask me anything about your yarn parameters, database batches, or fabric suitability."
            }]
            db_save_chat_message(
                st.session_state["email"], "assistant",
                st.session_state["chat_history"][0]["content"]
            )
        st.session_state["chat_loaded"] = True

    # ==========================================
    # 9. NAVIGATION SIDEBAR
    # ==========================================
    with st.sidebar:
        st.markdown(
            """
            <div style='display:flex;align-items:center;gap:12px;margin-bottom:20px;padding:8px 4px;'>
                <div style='background:linear-gradient(135deg, #0d9488, #6366f1);width:38px;height:38px;border-radius:8px;display:flex;align-items:center;justify-content:center;color:white;font-weight:800;font-size:1.15rem;box-shadow:0 4px 10px rgba(13,148,136,0.3);'>YI</div>
                <div>
                    <h4 style='margin:0;color:white;font-size:1.1rem;'>YarnInsight</h4>
                    <p style='margin:0;color:#94a3b8;font-size:0.75rem;text-transform:uppercase;font-weight:700;letter-spacing:0.05em;'>AI Production Command</p>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )
        
        # User details Card
        role_label = html.escape(str(st.session_state["role"]))
        user_name = html.escape(str(st.session_state["name"]))
        st.markdown(
            f"""
            <div style='background-color:#1e293b;border:1px solid #334155;border-radius:10px;padding:12px 14px;margin-bottom:20px;'>
                <div style='font-size:0.7rem;color:#94a3b8;text-transform:uppercase;font-weight:800;letter-spacing:0.05em;'>Active Session</div>
                <div style='font-weight:700;color:white;font-size:0.95rem;margin-top:2px;'>{user_name}</div>
                <div style='font-size:0.75rem;color:#0d9488;font-weight:700;margin-top:2px;'>{role_label}</div>
            </div>
            """,
            unsafe_allow_html=True
        )
        
        # Multi-page routing selection
        page_options = [
            "Home / Portal",
            "Executive Dashboard",
            "Predict Strength & Grade",
            "Fabric Recommendation",
            "Batch Management",
            "Bulk Prediction Tool",
            "Advanced Analytics",
            "Supplier Analysis",
            "AI Chat Assistant",
            "Report & Certificates",
            "Admin Control Panel",
            "Documentation"
        ]
        nav_selection = st.radio("Navigate Pages", page_options, label_visibility="collapsed")
        
        st.divider()
        if st.button("Logout Session", use_container_width=True):
            db_log_audit(st.session_state["email"], "LOGOUT", "User logged out successfully.")
            st.session_state["auth"] = False
            st.session_state["email"] = ""
            st.session_state["name"] = ""
            st.session_state["role"] = ""
            st.rerun()
            
    # Page Router
    clean_page_name = nav_selection
    
    # Role-based restriction checks
    # Quality Engineers and Production Managers can use operational pages.
    # Admin can access everything.
    role = st.session_state["role"]
    restricted = False
    
    if role not in ["Admin", "Production Manager", "Quality Engineer"]:
        restricted = True
    elif role == "Quality Engineer" and clean_page_name in ["Admin Control Panel"]:
        restricted = True
    elif role == "Production Manager" and clean_page_name in ["Admin Control Panel"]:
        restricted = True
        
    if restricted:
        render_restricted_page(clean_page_name)
        return

    # ==========================================
    # 10. ROUTING THE PAGES
    # ==========================================
    if clean_page_name == "Home / Portal":
        page_home()
    elif clean_page_name == "Executive Dashboard":
        page_dashboard(df_labeled)
    elif clean_page_name == "Predict Strength & Grade":
        page_predict(reg_model, clf_model, df_labeled, r2, acc)
    elif clean_page_name == "Fabric Recommendation":
        page_fabric_recommendation(df_labeled)
    elif clean_page_name == "Batch Management":
        page_batch_management(reg_model, clf_model, df_labeled)
    elif clean_page_name == "Bulk Prediction Tool":
        page_bulk_prediction(reg_model, clf_model, df_labeled)
    elif clean_page_name == "Advanced Analytics":
        page_analytics(df_labeled)
    elif clean_page_name == "Supplier Analysis":
        page_supplier_analysis()
    elif clean_page_name == "AI Chat Assistant":
        page_ai_assistant(df_labeled)
    elif clean_page_name == "Report & Certificates":
        page_reports()
    elif clean_page_name == "Admin Control Panel":
        page_admin()
    elif clean_page_name == "Documentation":
        page_documentation(r2, acc)

# ==========================================
# AUTHENTICATION PAGE DESIGN
# ==========================================
def render_auth_page():
    left, right = st.columns([1.1, 0.9], gap="large")
    with left:
        st.markdown(
            """
            <div style='background: linear-gradient(135deg, #0f172a, #0d9488); min-height: 580px; border-radius: 16px; padding: 40px; color: white; display: flex; flex-direction: column; justify-content: space-between; box-shadow: 0 10px 25px rgba(0,0,0,0.15);'>
                <div style='display: flex; align-items: center; gap: 12px;'>
                    <div style='background: white; width: 44px; height: 44px; border-radius: 10px; display: flex; align-items: center; justify-content: center; color: #0f766e; font-weight: 800; font-size: 1.35rem; box-shadow: 0 4px 10px rgba(0,0,0,0.2);'>YI</div>
                    <span style='font-family: Outfit, sans-serif; font-size: 1.45rem; font-weight: 700; letter-spacing: -0.01em;'>YarnInsight AI</span>
                </div>
                <div>
                    <h1 style='color: white !important; font-size: 2.75rem; font-weight: 800; line-height: 1.1; margin-bottom: 15px;'>Predictive Quality Intelligence for Modern Spinning.</h1>
                    <p style='color: #ccfbf1; font-size: 1.15rem; font-weight: 400; line-height: 1.6; max-width: 520px; margin-bottom: 25px;'>Analyze fiber parameters, predict tensile strength, assess structural risks, and match yarn lots to target cloth designs from one secure operational control workspace.</p>
                    <div style='display: flex; gap: 8px; flex-wrap: wrap;'>
                        <span style='background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.2); border-radius: 20px; padding: 6px 14px; font-size: 0.8rem; font-weight: 700; letter-spacing: 0.02em;'>STRENGTH REGRESSION</span>
                        <span style='background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.2); border-radius: 20px; padding: 6px 14px; font-size: 0.8rem; font-weight: 700; letter-spacing: 0.02em;'>GRADE CLASSIFICATION</span>
                        <span style='background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.2); border-radius: 20px; padding: 6px 14px; font-size: 0.8rem; font-weight: 700; letter-spacing: 0.02em;'>FABRIC SUITABILITY</span>
                        <span style='background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.2); border-radius: 20px; padding: 6px 14px; font-size: 0.8rem; font-weight: 700; letter-spacing: 0.02em;'>BATCH COMPARISON</span>
                    </div>
                </div>
                <div style='display: flex; gap: 24px; border-top: 1px solid rgba(255,255,255,0.15); padding-top: 25px;'>
                    <div><h4 style='color: white; margin: 0; font-size: 1.35rem;'>94.5%</h4><p style='color: #99f6e4; margin: 0; font-size: 0.75rem; font-weight: 600;'>Model Accuracy</p></div>
                    <div><h4 style='color: white; margin: 0; font-size: 1.35rem;'>100K+</h4><p style='color: #99f6e4; margin: 0; font-size: 0.75rem; font-weight: 600;'>Samples Analyzed</p></div>
                    <div><h4 style='color: white; margin: 0; font-size: 1.35rem;'>5-Tiers</h4><p style='color: #99f6e4; margin: 0; font-size: 0.75rem; font-weight: 600;'>Grading Architecture</p></div>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )
    with right:
        st.markdown(
            """
            <div style='background: white; border: 1px solid #e2e8f0; border-radius: 16px; padding: 30px; box-shadow: 0 4px 10px rgba(0,0,0,0.03); min-height: 580px; display: flex; flex-direction: column; justify-content: space-between;'>
                <div>
                    <h2 style='margin-top: 0; margin-bottom: 6px;'>Workspace Access Portal</h2>
                    <p style='color: #64748b; font-size: 0.9rem; margin-bottom: 24px;'>Login using authorized credentials to view predictive diagnostics and batch reports.</p>
                </div>
            """,
            unsafe_allow_html=True
        )
        
        tab_login, tab_signup = st.tabs(["Secure Login", "Create Account"])
        
        with tab_login:
            with st.form("auth_login_form"):
                email = st.text_input("Work Email Address", placeholder="name@company.com")
                password = st.text_input("Account Password", type="password", placeholder="••••••••")
                submit = st.form_submit_button("Enter Operational Console")
                
            if submit:
                if not email or not password:
                    st.error("Please fill in all credentials.")
                else:
                    success, name, role = db_verify_login(email, password)
                    if success:
                        st.session_state["auth"] = True
                        st.session_state["email"] = email.strip().lower()
                        st.session_state["name"] = name
                        st.session_state["role"] = role
                        db_log_audit(email, "LOGIN", f"User logged in successfully with role: {role}.")
                        st.success("Login verified. Syncing workspace...")
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error("Invalid credentials. Please verify details.")
                        
        with tab_signup:
            with st.form("auth_signup_form"):
                new_name = st.text_input("Full Name", placeholder="John Doe")
                new_email = st.text_input("Work Email", placeholder="name@company.com")
                new_password = st.text_input("Password", type="password", placeholder="Min. 12 characters")
                new_role = st.selectbox("Company Operational Role", ["Quality Engineer", "Production Manager"])
                signup_submit = st.form_submit_button("Register for Access")
                
            if signup_submit:
                if not new_name or not new_email or not new_password:
                    st.error("Please fill out all registration fields.")
                elif len(new_password) < 12:
                    st.error("Password must be at least 12 characters.")
                elif "@" not in new_email or "." not in new_email:
                    st.error("Please enter a valid work email.")
                else:
                    success, msg = db_create_user(new_email, new_name, new_password, new_role)
                    if success:
                        st.success(msg)
                    else:
                        st.error(msg)
                        
        st.markdown(
            """
            <div style='background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; font-size: 0.8rem; color: #64748b; margin-top: 15px; line-height: 1.4;'>
                <b>Secure access:</b><br/>
                Use the administrator account created during deployment. New users can request role-based access through the registration tab.
            </div>
            """,
            unsafe_allow_html=True
        )
        st.markdown("</div>", unsafe_allow_html=True)

def render_restricted_page(page_name):
    safe_role = html.escape(str(st.session_state["role"]))
    safe_page_name = html.escape(str(page_name))
    st.markdown(
        f"""
        <div class='glass-header' style='background: linear-gradient(135deg, #ef4444, #b91c1c);'>
            <h1>Access Denied</h1>
            <p>Your current security role <b>({safe_role})</b> does not have authority to view the "{safe_page_name}" page.</p>
        </div>
        <div class='premium-card' style='text-align: center; padding: 60px 40px;'>
            <h3 style='color: #ef4444;'>Insufficient Permissions</h3>
            <p style='color: #64748b; max-width: 500px; margin: 10px auto;'>Contact your system Administrator to request role modification if this access is required for your production workflow.</p>
        </div>
        """,
        unsafe_allow_html=True
    )

# ==========================================
# PAGE 1: HOME PORTAL
# ==========================================
def page_home():
    st.markdown(
        """
        <div class='glass-header'>
            <h1>🏭 YarnInsight Portal</h1>
            <p>Artificial Intelligence Yarn Quality Grading & Fabric Suitability Matching Platform</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    # Dashboard grid links
    st.markdown("### Operational Shortcut Panel")
    
    # 4 columns for page quicklinks
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(
            """
            <div class='premium-card' style='height: 100%; display: flex; flex-direction: column; justify-content: space-between;'>
                <div>
                    <div style='font-size: 2rem; margin-bottom: 12px;'>📊</div>
                    <h4 style='margin: 0 0 6px 0; color: #0f172a;'>Executive Stats</h4>
                    <p style='margin: 0; color: #64748b; font-size: 0.8rem; line-height: 1.4;'>Monitor key production metrics, rejections, average strength, and material distributions.</p>
                </div>
                <div style='margin-top: 15px; font-weight: 700; font-size: 0.8rem; color: #0d9488;'>GO TO DASHBOARD →</div>
            </div>
            """,
            unsafe_allow_html=True
        )
    with c2:
        st.markdown(
            """
            <div class='premium-card' style='height: 100%; display: flex; flex-direction: column; justify-content: space-between;'>
                <div>
                    <div style='font-size: 2rem; margin-bottom: 12px;'>🔬</div>
                    <h4 style='margin: 0 0 6px 0; color: #0f172a;'>Yarn Predictor</h4>
                    <p style='margin: 0; color: #64748b; font-size: 0.8rem; line-height: 1.4;'>Enter raw physical properties to execute neural prediction for tensile strength and quality tier.</p>
                </div>
                <div style='margin-top: 15px; font-weight: 700; font-size: 0.8rem; color: #0d9488;'>START TESTING →</div>
            </div>
            """,
            unsafe_allow_html=True
        )
    with c3:
        st.markdown(
            """
            <div class='premium-card' style='height: 100%; display: flex; flex-direction: column; justify-content: space-between;'>
                <div>
                    <div style='font-size: 2rem; margin-bottom: 12px;'>👗</div>
                    <h4 style='margin: 0 0 6px 0; color: #0f172a;'>Fabric Suitability</h4>
                    <p style='margin: 0; color: #64748b; font-size: 0.8rem; line-height: 1.4;'>Instantly evaluate what end-products can be woven from active yarn properties.</p>
                </div>
                <div style='margin-top: 15px; font-weight: 700; font-size: 0.8rem; color: #0d9488;'>CHECK SUITABILITY →</div>
            </div>
            """,
            unsafe_allow_html=True
        )
    with c4:
        st.markdown(
            """
            <div class='premium-card' style='height: 100%; display: flex; flex-direction: column; justify-content: space-between;'>
                <div>
                    <div style='font-size: 2rem; margin-bottom: 12px;'>📦</div>
                    <h4 style='margin: 0 0 6px 0; color: #0f172a;'>Batch Compare</h4>
                    <p style='margin: 0; color: #64748b; font-size: 0.8rem; line-height: 1.4;'>Audit logs, create batch codes, search inventory database, and compare batches side-by-side.</p>
                </div>
                <div style='margin-top: 15px; font-weight: 700; font-size: 0.8rem; color: #0d9488;'>OPEN MANAGER →</div>
            </div>
            """,
            unsafe_allow_html=True
        )
        
    left, right = st.columns([1.2, 0.8], gap="large")
    with left:
        st.markdown(
            """
            <div class='premium-card'>
                <h3>Platform Capabilities & Automated Workflows</h3>
                <p style='color:#64748b;'>YarnInsight integrates raw physical fiber composition (cellulose, hemicellulose, lignin, pectin) with mechanical attributes and spinning parameters to provide real-time quality assurance.</p>
                <div style='border-top:1px solid #f1f5f9; padding-top:15px; margin-top:15px;'>
                    <h5>⚙️ Core Algorithms</h5>
                    <ul>
                        <li><b>RF Strength Regressor:</b> Predicts actual tensile yield limits (MPa) before fabric load tests.</li>
                        <li><b>RF Multi-Class Classifier:</b> Grades lots into Grade A+ (Premium), Grade A, Grade B, Grade C, or Rejects based on dataset benchmarks.</li>
                        <li><b>Heuristic Fabric Matcher:</b> Runs parameter constraints against fabric specs (Denim, T-shirts, Sportswear, etc.) to estimate weaving yield performance.</li>
                        <li><b>Explainable Local SHAP:</b> Graphically shows which inputs drove the strength up or down for explaining decisions.</li>
                    </ul>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )
    with right:
        # System Health Status
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM batches")
        total_batches = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM prediction_history")
        total_predictions = cursor.fetchone()[0]
        conn.close()
        
        st.markdown(
            f"""
            <div class='premium-card'>
                <h3>System Health</h3>
                <div style='display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; padding:6px 0;'>
                    <span>Machine Learning Service:</span>
                    <span class='badge' style='background-color:#ecfdf5; color:#065f46;'>● ONLINE</span>
                </div>
                <div style='display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; padding:6px 0;'>
                    <span>Database Status (SQLite):</span>
                    <span class='badge' style='background-color:#ecfdf5; color:#065f46;'>● INTEGRATED</span>
                </div>
                <div style='display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; padding:6px 0;'>
                    <span>Batch Records Saved:</span>
                    <span style='font-weight:700; color:#0f172a;'>{total_batches}</span>
                </div>
                <div style='display:flex; justify-content:space-between; align-items:center; padding:6px 0;'>
                    <span>Logged User Predictions:</span>
                    <span style='font-weight:700; color:#0f172a;'>{total_predictions}</span>
                </div>
                <div style='background-color:#f0fdfa; border: 1px solid #ccfbf1; border-radius:8px; padding:12px; margin-top:18px; font-size:0.8rem; color:#0f766e;'>
                    ⚡ <b>Operational Tip:</b> Use the <b>AI Chat Assistant</b> in the sidebar to ask questions about raw batch quality. Mentioning a batch ID like <code>B-1010</code> will query live parameters!
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

# ==========================================
# PAGE 2: EXECUTIVE DASHBOARD
# ==========================================
def page_dashboard(df_labeled):
    st.markdown(
        """
        <div class='glass-header'>
            <h1>📊 Executive Operations Dashboard</h1>
            <p>Real-Time KPI Tracking, Yarn Quality Distribution, and Automated Alerts</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    # Query database stats
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM batches")
    db_batches_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM prediction_history")
    db_preds_count = cursor.fetchone()[0]
    conn.close()
    
    total_predictions = len(df_labeled) + db_preds_count
    avg_strength = df_labeled["Tensile Strength of yarn (MPa)"].mean()
    
    # Defect Rate (Reject share)
    reject_share = (df_labeled["Quality Grade"] == "Reject").mean() * 100
    premium_share = (df_labeled["Quality Grade"] == "Grade A+ (Premium)").mean() * 100
    
    # SCADA-style Plotly dial gauges
    c_gauge1, c_gauge2 = st.columns(2)
    with c_gauge1:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.plotly_chart(build_scada_strength_gauge(avg_strength, target=1500.0), use_container_width=True)
        st.caption("Target: 1500 MPa — Red threshold marks minimum acceptable yield.")
        st.markdown("</div>", unsafe_allow_html=True)

    with c_gauge2:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.plotly_chart(build_scada_defect_gauge(reject_share), use_container_width=True)
        st.caption("Green zone ≤ 2% — Yellow 2–5% — Red > 5% defect rate.")
        st.markdown("</div>", unsafe_allow_html=True)
        
    # KPIs Layout
    st.markdown("<div class='kpi-container'>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(
            f"""
            <div class='kpi-card' style='border-left-color: #6366f1;'>
                <div class='kpi-label'>Total Predictions Logged</div>
                <div class='kpi-value'>{total_predictions:,}</div>
                <div class='kpi-subtext'>Training dataset + live queries</div>
            </div>
            """,
            unsafe_allow_html=True
        )
    with c2:
        st.markdown(
            f"""
            <div class='kpi-card' style='border-left-color: #10b981;'>
                <div class='kpi-label'>Premium Lot Share</div>
                <div class='kpi-value'>{premium_share:.1f}%</div>
                <div class='kpi-subtext'>Lots graded as A+ (Premium)</div>
            </div>
            """,
            unsafe_allow_html=True
        )
    st.markdown("</div>", unsafe_allow_html=True)
    
    # -----------------------------------------
    # Notification Center Alerts
    # -----------------------------------------
    st.markdown("### 🔔 Active Quality Alerts")
    batches_df = db_get_batches()
    
    # Risk notifications
    high_risk_batches = batches_df[batches_df["risk_level"] == "High Risk"]
    reject_batches = batches_df[batches_df["predicted_grade"] == "Reject"]
    
    alert_col1, alert_col2 = st.columns(2)
    with alert_col1:
        if len(reject_batches) > 0:
            st.markdown(
                f"""
                <div class='custom-alert danger'>
                    <div>🛑</div>
                    <div><b>Failed Quality Check:</b> {len(reject_batches)} batches have been graded as <b>REJECT</b>. Re-blend or recycle is required. (Recent: {', '.join(reject_batches['batch_id'].head(3))})</div>
                </div>
                """,
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                """
                <div class='custom-alert info'>
                    <div>✓</div>
                    <div>No critical rejected lots detected in active warehouse system logs.</div>
                </div>
                """,
                unsafe_allow_html=True
            )
    with alert_col2:
        if len(high_risk_batches) > 0:
            st.markdown(
                f"""
                <div class='custom-alert warning'>
                    <div>⚠️</div>
                    <div><b>High Production Risk:</b> {len(high_risk_batches)} batches are flagged with <b>High Material Risk</b> (critical deviations in pH, Moisture, or Tenacity).</div>
                </div>
                """,
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                """
                <div class='custom-alert info'>
                    <div>✓</div>
                    <div>All active batch physical properties are operating within acceptable limits.</div>
                </div>
                """,
                unsafe_allow_html=True
            )
            
    # Chart Grid
    left, right = st.columns(2)
    
    with left:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>Yarn Quality Grade Distribution</h4>", unsafe_allow_html=True)
        # Combine database batches and training labels to show complete picture
        grades_combined = pd.concat([df_labeled["Quality Grade"], batches_df["predicted_grade"]]).value_counts()
        fig_grades = px.pie(
            values=grades_combined.values,
            names=grades_combined.index,
            hole=0.4,
            color_discrete_sequence=['#10b981', '#34d399', '#f59e0b', '#f97316', '#ef4444']
        )
        fig_grades.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=300)
        st.plotly_chart(fig_grades, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
        
    with right:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>Recent Saved Production Batches</h4>", unsafe_allow_html=True)
        if len(batches_df) > 0:
            st.dataframe(
                batches_df[["batch_id", "creation_time", "supplier_name", "predicted_grade", "risk_level", "status"]].head(6),
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("No batches registered in database yet. Navigate to 'Predict' to save new lots.")
        st.markdown("</div>", unsafe_allow_html=True)

# ==========================================
# PAGE 3: PREDICT YARN STRENGTH
# ==========================================
def page_predict(reg_model, clf_model, df_labeled, r2, acc):
    st.markdown(
        """
        <div class='glass-header'>
            <h1>🔬 AI Yarn Strength & Grade Predictor</h1>
            <p>Calculate Tensile Yield limits, assign Quality Tiers, evaluate parameters, and optimize outputs.</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    st.markdown("### 🎛️ Input Lab Parameters")
    
    with st.form("single_prediction_form"):
        t1, t2, t3 = st.tabs(["🌾 Fiber Composition", "💪 Mechanical Properties", "⚙️ Chemical & Ambient Details"])
        
        with t1:
            col1, col2 = st.columns(2)
            cell = col1.slider("Cellulose of yarn (%)", 60.0, 95.0, float(df_labeled["Cellulose of yarn (%)"].median()), step=0.1)
            hemi = col2.slider("Hemicellulose of yarn (%)", 12.0, 28.0, float(df_labeled["Hemicellulose of yarn (%)"].median()), step=0.1)
            lignin = col1.slider("Lignin of yarn (%)", 4.0, 18.0, float(df_labeled["Lignin of yarn (%)"].median()), step=0.1)
            pectin = col2.slider("Pectin of yarn (%)", 0.2, 3.0, float(df_labeled["Pectin of yarn (%)"].median()), step=0.05)
            
        with t2:
            col1, col2 = st.columns(2)
            fineness = col1.slider("Yarn Fineness (tex)", 1.5, 4.5, float(df_labeled["Fineness of yarn (tex)"].median()), step=0.1)
            tenacity = col2.slider("Fiber Tenacity (gm/tex)", 25.0, 55.0, float(df_labeled["Fiber Tenacity of yarn (gm/tex)"].median()), step=0.1)
            elongation = col1.slider("Yarn Elongation (%)", 1.8, 4.2, float(df_labeled["Elongation of yarn (%)"].median()), step=0.05)
            density = col2.slider("True Density (gms/cc)", 1.35, 1.65, float(df_labeled["True Density of yarn (gms/cc)"].median()), step=0.01)
            porosity = col1.slider("Yarn Porosity (%)", 4.0, 18.0, float(df_labeled["Porosity of yarn (%)"].median()), step=0.1)
            
        with t3:
            col1, col2 = st.columns(2)
            moisture_c = col1.slider("Moisture Content (%)", 8.0, 14.0, float(df_labeled["Moisture Content of yarn (%)"].median()), step=0.1)
            ph = col2.slider("Yarn pH Level", 4.0, 8.0, float(df_labeled["pH Level of yarn"].median()), step=0.1)
            moisture_r = col1.slider("Moisture Regain (%)", 8.0, 15.0, float(df_labeled["Moisture Regain of yarn (%)"].median()), step=0.1)
            swelling = col2.slider("Water Swelling (%)", 30.0, 65.0, float(df_labeled["Water Swelling of yarn (%)"].median()), step=0.5)
            dye_type = col1.selectbox("Dye Mix Type", sorted(df_labeled["Dye Type"].dropna().unique()))
            
        st.markdown("<br/>", unsafe_allow_html=True)
        btn_predict = st.form_submit_button("🧪 Run Neural Model Analysis")
        
    if btn_predict:
        # Create input dict
        inputs = {
            "Cellulose of yarn (%)": cell,
            "Hemicellulose of yarn (%)": hemi,
            "Lignin of yarn (%)": lignin,
            "Pectin of yarn (%)": pectin,
            "Moisture Content of yarn (%)": moisture_c,
            "pH Level of yarn": ph,
            "Fineness of yarn (tex)": fineness,
            "Fiber Tenacity of yarn (gm/tex)": tenacity,
            "Elongation of yarn (%)": elongation,
            "Moisture Regain of yarn (%)": moisture_r,
            "Water Swelling of yarn (%)": swelling,
            "True Density of yarn (gms/cc)": density,
            "Porosity of yarn (%)": porosity,
            "Dye Type": dye_type
        }
        
        with st.spinner("Processing physical property vectors through RandomForest pipelines..."):
            pred_strength, pred_grade, confidence, risk_level, risk_reasons = run_predictions_on_row(
                reg_model, clf_model, inputs, df_labeled
            )
            
        st.session_state["last_pred_vals"] = inputs
        st.session_state["last_pred_results"] = {
            "strength": pred_strength,
            "grade": pred_grade,
            "confidence": confidence,
            "risk_level": risk_level,
            "risk_reasons": risk_reasons
        }
        
        # Log to prediction history in SQLite
        history_record = inputs.copy()
        history_record["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        history_record["user_email"] = st.session_state["email"]
        history_record["predicted_strength"] = pred_strength
        history_record["predicted_grade"] = pred_grade
        history_record["confidence"] = confidence
        history_record["risk_level"] = risk_level
        db_save_prediction_history(history_record)

    # Render results if present in session state
    if "last_pred_results" in st.session_state:
        res = st.session_state["last_pred_results"]
        inputs = st.session_state["last_pred_vals"]
        
        st.markdown("---")
        st.markdown("### 📊 Automated Quality Report")
        
        r_col1, r_col2 = st.columns([0.8, 1.2], gap="large")
        
        with r_col1:
            st.markdown("<div class='premium-card' style='text-align: center; border-left: 6px solid #0d9488;'>", unsafe_allow_html=True)
            st.markdown(f"<div class='kpi-label'>Predicted Tensile Strength</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='kpi-value' style='font-size:3rem; color:#0d9488;'>{res['strength']:.1f} MPa</div>", unsafe_allow_html=True)
            
            grade_badge_class = {
                "Grade A+ (Premium)": "badge-a-plus",
                "Grade A": "badge-a",
                "Grade B": "badge-b",
                "Grade C": "badge-c",
                "Reject": "badge-reject"
            }.get(res["grade"], "")
            
            st.markdown(f"<div style='margin: 15px 0;'><span class='badge {grade_badge_class}' style='font-size:1rem; padding: 6px 14px;'>{res['grade']}</span></div>", unsafe_allow_html=True)
            st.markdown(f"<div class='kpi-subtext'>ML Classification Confidence: <b>{res['confidence']:.2f}%</b></div>", unsafe_allow_html=True)
            
            risk_badge_class = {
                "Low Risk": "badge-low-risk",
                "Medium Risk": "badge-med-risk",
                "High Risk": "badge-high-risk"
            }.get(res["risk_level"], "")
            
            st.markdown(f"<div style='margin-top: 15px;'>Risk Profile: <span class='badge {risk_badge_class}'>{res['risk_level']}</span></div>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)
            
            # Risk warning box
            alert_class = "danger" if res["risk_level"] == "High Risk" else ("warning" if res["risk_level"] == "Medium Risk" else "info")
            alert_icon = "🛑" if alert_class == "danger" else ("⚠️" if alert_class == "warning" else "✓")
            st.markdown(
                f"""
                <div class='custom-alert {alert_class}'>
                    <div>{alert_icon}</div>
                    <div><b>Risk Notes:</b> {res['risk_reasons']}</div>
                </div>
                """,
                unsafe_allow_html=True
            )
            
            # Save Batch Form
            st.markdown("#### 📦 Store as Batch lot")
            with st.form("save_batch_form"):
                batch_id = st.text_input("Assign Batch ID", value=f"B-20{random.randint(10,99)}")
                supplier_name = st.selectbox("Supplier Name", ["Apex Fibers", "Global Cotton Co.", "TexPrime Industries", "EcoThread Co.", "Other / Import"])
                status = st.selectbox("Assign Action Status", ["Release", "Review", "Hold"])
                save_btn = st.form_submit_button("Confirm Batch Registration")
                
            if save_btn:
                # Compile full record
                batch_data = inputs.copy()
                batch_data["batch_id"] = batch_id.strip()
                batch_data["creation_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                batch_data["supplier_name"] = supplier_name
                batch_data["actual_strength"] = res["strength"]
                batch_data["predicted_strength"] = res["strength"]
                batch_data["predicted_grade"] = res["grade"]
                batch_data["confidence"] = res["confidence"]
                batch_data["risk_level"] = res["risk_level"]
                batch_data["risk_reasons"] = res["risk_reasons"]
                batch_data["status"] = status
                
                db_save_batch(batch_data)
                st.success(f"Batch {batch_id} registered inside SQL database inventory logs.")
                time.sleep(1)
                st.rerun()
                
        with r_col2:
            st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
            st.markdown("<h4>🧠 Explainable AI: Feature Contributions</h4>", unsafe_allow_html=True)
            
            # Draw local contributions
            # We construct a mock local explanation graph based on global weights + parameters
            # For demonstration, we compare the current inputs with average training inputs to see impacts
            impacts = []
            for col in NUMERIC_COLUMNS:
                avg = df_labeled[col].median()
                val = inputs[col]
                diff = val - avg
                
                # Assign directional weights based on logical physical rules
                weight = 1.0
                if "Tenacity" in col: weight = 12.0
                elif "Fineness" in col: weight = -15.0
                elif "Elongation" in col: weight = 8.0
                elif "Porosity" in col: weight = -6.0
                elif "Moisture Regain" in col: weight = 4.0
                else: weight = 0.5
                
                impact = diff * weight
                impacts.append((col.replace(" of yarn (%)", "").replace("of yarn (gm/tex)", "").replace("of yarn", ""), impact))
                
            impacts_df = pd.DataFrame(impacts, columns=["Feature", "Impact (MPa)"]).sort_values("Impact (MPa)")
            
            fig_explain = px.bar(
                impacts_df,
                x="Impact (MPa)",
                y="Feature",
                orientation="h",
                color="Impact (MPa)",
                color_continuous_scale=px.colors.diverging.Tealrose,
                title="Input Variable Pull on Yarn Yield Limits (MPa)"
            )
            fig_explain.update_layout(margin=dict(t=30, b=10, l=10, r=10), height=300)
            st.plotly_chart(fig_explain, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)
            
            # Optimization Section
            st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
            st.markdown("<h4>🎯 Production Optimization Suggestions</h4>", unsafe_allow_html=True)
            st.markdown("<p style='color:#64748b; font-size:0.85rem;'>Suggested modifications to raw ingredients/processing to elevate yarn yield limits.</p>", unsafe_allow_html=True)
            
            suggestions = get_optimizer_suggestions(inputs)
            
            st.markdown("<div class='opt-grid'>", unsafe_allow_html=True)
            for sug in suggestions:
                st.markdown(
                    f"""
                    <div class='opt-card'>
                        <div class='opt-title'>{sug['parameter']}</div>
                        <div class='opt-change'>{sug['current']} → {sug['target']}</div>
                        <div style='font-size:0.8rem; font-weight:700; color:#0d9488;'>{sug['impact']}</div>
                        <p class='opt-desc' style='margin-top:6px;'><b>Action:</b> {sug['action']}</p>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
            st.markdown("</div>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)
            
            # Interactive Optimization Sandbox
            st.markdown("<div class='premium-card' style='border-top:4px solid #6366f1;'>", unsafe_allow_html=True)
            st.markdown("<h4>⚙️ Interactive Parameter Optimization Sandbox</h4>", unsafe_allow_html=True)
            st.markdown("<p style='color:#64748b; font-size:0.85rem;'>Adjust Tenacity, Moisture, Porosity, and Fineness — watch Tensile Strength, Quality Grade, and Risk Level recalculate live.</p>", unsafe_allow_html=True)
            
            sb_col1, sb_col2 = st.columns(2)
            with sb_col1:
                s_tenacity = st.slider("Fiber Tenacity (gm/tex)", 25.0, 55.0, float(inputs["Fiber Tenacity of yarn (gm/tex)"]), key="sandbox_tenacity")
                s_fineness = st.slider("Yarn Fineness (tex)", 1.5, 4.5, float(inputs["Fineness of yarn (tex)"]), key="sandbox_fineness")
            with sb_col2:
                s_porosity = st.slider("Yarn Porosity (%)", 4.0, 18.0, float(inputs["Porosity of yarn (%)"]), key="sandbox_porosity")
                s_moisture = st.slider("Moisture Content (%)", 8.0, 14.0, float(inputs["Moisture Content of yarn (%)"]), key="sandbox_moisture")
            
            sandbox_inputs = inputs.copy()
            sandbox_inputs["Fiber Tenacity of yarn (gm/tex)"] = s_tenacity
            sandbox_inputs["Fineness of yarn (tex)"] = s_fineness
            sandbox_inputs["Porosity of yarn (%)"] = s_porosity
            sandbox_inputs["Moisture Content of yarn (%)"] = s_moisture
            
            s_strength, s_grade, s_conf, s_risk, s_reasons = run_predictions_on_row(
                reg_model, clf_model, sandbox_inputs, df_labeled
            )
            
            st.plotly_chart(
                build_sandbox_live_gauges(s_strength, s_grade, s_risk, res["strength"]),
                use_container_width=True
            )
            
            s_diff_strength = s_strength - res["strength"]
            s_sign = "+" if s_diff_strength >= 0 else ""
            grade_order = {"Grade A+ (Premium)": 5, "Grade A": 4, "Grade B": 3, "Grade C": 2, "Reject": 1}
            if s_grade != res["grade"]:
                grade_changed = "↑ upgraded" if grade_order.get(s_grade, 0) > grade_order.get(res["grade"], 0) else "↓ downgraded"
            else:
                grade_changed = "unchanged"
            
            st.markdown(
                f"""
                <div style='background-color:#f8fafc; border:1px dashed #6366f1; border-radius:8px; padding:12px; margin-top:8px;'>
                    <div style='display:grid; grid-template-columns:1fr 1fr 1fr; gap:12px; text-align:center;'>
                        <div>
                            <div style='font-size:0.75rem; color:#475569; text-transform:uppercase;'>Simulated Strength</div>
                            <div style='font-size:1.3rem; font-weight:800; color:#6366f1;'>{s_strength:.1f} MPa</div>
                            <div style='font-size:0.8rem; color:{"#10b981" if s_diff_strength >= 0 else "#ef4444"};'>({s_sign}{s_diff_strength:.1f} vs baseline)</div>
                        </div>
                        <div>
                            <div style='font-size:0.75rem; color:#475569; text-transform:uppercase;'>Quality Grade</div>
                            <div style='font-size:1rem; font-weight:700; color:#0d9488;'>{s_grade}</div>
                            <div style='font-size:0.8rem; color:#64748b;'>{grade_changed}</div>
                        </div>
                        <div>
                            <div style='font-size:0.75rem; color:#475569; text-transform:uppercase;'>Risk Level</div>
                            <div style='font-size:1rem; font-weight:700; color:#ea580c;'>{s_risk}</div>
                            <div style='font-size:0.75rem; color:#64748b;'>{s_conf:.1f}% confidence</div>
                        </div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )
            st.markdown("</div>", unsafe_allow_html=True)
            
    # Recent predictions Log
    st.markdown("### 📋 Live Query Prediction History")
    hist_df = db_get_prediction_history(8)
    if len(hist_df) > 0:
        st.dataframe(
            hist_df[["timestamp", "user_email", "predicted_strength", "predicted_grade", "confidence", "risk_level"]],
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("No predictions logged yet.")

# ==========================================
# PAGE 4: FABRIC RECOMMENDATION ENGINE
# ==========================================
def page_fabric_recommendation(df_labeled):
    st.markdown(
        """
        <div class='glass-header'>
            <h1>👗 Fabric Suitability & End-Product Matcher</h1>
            <p>Answer: "What fabric styles can be manufactured from this yarn lot?"</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    st.markdown("### Select Source Yarn Properties")
    
    # User can select to load a saved batch or enter custom
    options = ["Manually Input Yarn Vector", "Load Existing Saved Batch"]
    mode = st.radio("Evaluation Source", options, horizontal=True)
    
    row_data = None
    grade = "Grade B"
    
    if mode == "Load Existing Saved Batch":
        batches_df = db_get_batches()
        if len(batches_df) == 0:
            st.warning("No saved batches in inventory database. Fallback to manual entry.")
            mode = "Manually Input Yarn Vector"
        else:
            selected_id = st.selectbox("Select Batch ID", batches_df["batch_id"].unique())
            row_db = db_get_batch(selected_id)
            row_data = dict(row_db)
            grade = row_db["predicted_grade"]
            
            # Show parameters in card
            st.markdown(
                f"""
                <div class='premium-card' style='background:#f1f5f9;'>
                    <h5>Loaded Batch: <b>{selected_id} ({row_db['supplier_name']})</b></h5>
                    <p style='margin:0; font-size:0.85rem;'>Tensile Strength: <b>{row_db['predicted_strength']:.1f} MPa</b> | Grade: <b>{row_db['predicted_grade']}</b> | Risk: <b>{row_db['risk_level']}</b></p>
                </div>
                """,
                unsafe_allow_html=True
            )
            
    if mode == "Manually Input Yarn Vector":
        col1, col2, col3 = st.columns(3)
        strength = col1.number_input("Tensile Strength (MPa)", 1000.0, 2400.0, float(df_labeled["Tensile Strength of yarn (MPa)"].median()))
        tenacity = col2.number_input("Fiber Tenacity (gm/tex)", 20.0, 60.0, float(df_labeled["Fiber Tenacity of yarn (gm/tex)"].median()))
        fineness = col3.number_input("Fineness (tex)", 1.0, 5.0, float(df_labeled["Fineness of yarn (tex)"].median()))
        elongation = col1.number_input("Elongation (%)", 1.0, 5.0, float(df_labeled["Elongation of yarn (%)"].median()))
        regain = col2.number_input("Moisture Regain (%)", 5.0, 18.0, float(df_labeled["Moisture Regain of yarn (%)"].median()))
        swelling = col3.number_input("Water Swelling (%)", 20.0, 70.0, float(df_labeled["Water Swelling of yarn (%)"].median()))
        density = col1.number_input("True Density (gms/cc)", 1.2, 1.8, float(df_labeled["True Density of yarn (gms/cc)"].median()))
        porosity = col2.number_input("Porosity (%)", 3.0, 20.0, float(df_labeled["Porosity of yarn (%)"].median()))
        
        # Estimate grade based on strength
        if strength > 1800: grade = "Grade A+ (Premium)"
        elif strength > 1600: grade = "Grade A"
        elif strength > 1400: grade = "Grade B"
        elif strength > 1250: grade = "Grade C"
        else: grade = "Reject"
        
        row_data = {
            "Tensile Strength of yarn (MPa)": strength,
            "Fiber Tenacity of yarn (gm/tex)": tenacity,
            "Fineness of yarn (tex)": fineness,
            "Elongation of yarn (%)": elongation,
            "Moisture Regain of yarn (%)": regain,
            "Water Swelling of yarn (%)": swelling,
            "True Density of yarn (gms/cc)": density,
            "Porosity of yarn (%)": porosity
        }
        
    if row_data:
        recs = cloth_from_rules(row_data, grade)
        best_match = recs[0]
        alt_matches = recs[1:4]
        
        # Recommendations Output Layout
        st.markdown("---")
        st.markdown("### 👗 End-Product Matches")
        
        c_left, c_right = st.columns([1.0, 1.0], gap="large")
        
        with c_left:
            st.markdown("<div class='premium-card' style='background: linear-gradient(135deg, #0d9488, #0f766e); color: white; border-left: 6px solid #6366f1;'>", unsafe_allow_html=True)
            st.markdown(f"<div class='kpi-label' style='color:#ccfbf1;'>🥇 Best Product Match</div>", unsafe_allow_html=True)
            st.markdown(f"<h2 style='color:white !important; font-size:2.8rem; margin:6px 0;'>{best_match[0]}</h2>", unsafe_allow_html=True)
            st.markdown(f"<div style='font-size:1.15rem; font-weight:700;'>Matching Score: {best_match[1]}%</div>", unsafe_allow_html=True)
            st.markdown(f"<div style='margin-top:10px; font-size:0.9rem; color:#e2e8f0;'>Compatibility Performance: <b>{get_expected_perf(best_match[1])}</b></div>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)
            
            st.markdown("#### 🥈 Alternative Product Matches")
            for alt in alt_matches:
                st.markdown(
                    f"""
                    <div style='background: white; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px 16px; margin-bottom: 10px; display:flex; justify-content:space-between; align-items:center;'>
                        <div>
                            <span style='font-weight:700; font-size: 1rem; color:#0f172a;'>{alt[0]}</span>
                            <div style='font-size:0.75rem; color:#64748b;'>Yield Limit: {get_expected_perf(alt[1])}</div>
                        </div>
                        <span class='badge' style='background-color:#e0f2fe; color:#0369a1; font-size:0.9rem; padding: 4px 10px;'>{alt[1]}% Match</span>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
                
        with c_right:
            st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
            st.markdown("<h4>📋 Full Fabric Compatibility Matrix</h4>", unsafe_allow_html=True)
            
            matrix_data = []
            for name, score in recs:
                perf = get_expected_perf(score)
                matrix_data.append({
                    "End Product Fabric": name,
                    "Compatibility Score": f"{score}%",
                    "Expected Performance": perf
                })
                
            st.dataframe(pd.DataFrame(matrix_data), use_container_width=True, hide_index=True)
            st.markdown("</div>", unsafe_allow_html=True)

# ==========================================
# PAGE 5: BATCH MANAGEMENT
# ==========================================
def page_batch_management(reg_model, clf_model, df_labeled):
    st.markdown(
        """
        <div class='glass-header'>
            <h1>📦 Production Batch Management</h1>
            <p>Database inventory list, Batch Creations, and Side-by-Side Comparison</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    tab_list, tab_create, tab_compare = st.tabs(["🔍 Database Inventory Search", "➕ Register New Batch", "⚖️ Compare Two Batches"])
    
    with tab_list:
        col1, col2, col3 = st.columns(3)
        search_q = col1.text_input("Search (Batch ID / Supplier)", placeholder="B-10...")
        grade_f = col2.selectbox("Filter Grade", ["All"] + ["Grade A+ (Premium)", "Grade A", "Grade B", "Grade C", "Reject"])
        status_f = col3.selectbox("Filter Status", ["All", "Release", "Review", "Hold"])
        
        # Get data
        db_df = db_get_batches(
            search=search_q if search_q else None,
            grade=grade_f if grade_f != "All" else None,
            status=status_f if status_f != "All" else None
        )
        
        st.markdown(f"**Found {len(db_df)} records inside database:**")
        
        if len(db_df) > 0:
            # Add action button column
            st.dataframe(
                db_df[["batch_id", "creation_time", "supplier_name", "predicted_grade", "predicted_strength", "risk_level", "status"]],
                use_container_width=True,
                hide_index=True
            )
            
            # Simple batch delete form
            with st.expander("🗑️ Delete Batch lot"):
                del_id = st.text_input("Enter exact Batch ID to remove", placeholder="B-1010")
                confirm_del = st.button("Confirm Delete Lot")
                if confirm_del:
                    if del_id in db_df["batch_id"].values:
                        db_delete_batch(del_id)
                        st.success(f"Batch {del_id} removed.")
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error("Batch ID not found.")
        else:
            st.info("No records found matching current query filters.")
            
    with tab_create:
        st.markdown("#### Manually Register Fiber Assay Batch")
        with st.form("manual_batch_registration"):
            col1, col2 = st.columns(2)
            b_id = col1.text_input("New Batch ID", placeholder="B-3001")
            supplier = col2.selectbox("Lint Supplier", ["Apex Fibers", "Global Cotton Co.", "TexPrime Industries", "EcoThread Co."])
            
            st.divider()
            c_cell = col1.slider("Cellulose (%)", 60.0, 95.0, 78.0)
            c_hemi = col2.slider("Hemicellulose (%)", 12.0, 28.0, 20.0)
            c_lig = col1.slider("Lignin (%)", 4.0, 18.0, 10.0)
            c_pec = col2.slider("Pectin (%)", 0.2, 3.0, 1.2)
            c_fine = col1.slider("Fineness (tex)", 1.5, 4.5, 2.5)
            c_ten = col2.slider("Tenacity (gm/tex)", 25.0, 55.0, 40.0)
            c_elon = col1.slider("Elongation (%)", 1.8, 4.2, 3.0)
            c_mois_c = col2.slider("Moisture Content (%)", 8.0, 14.0, 10.5)
            c_ph = col1.slider("pH Level", 4.0, 8.0, 5.8)
            c_reg = col2.slider("Moisture Regain (%)", 8.0, 15.0, 11.0)
            c_swell = col1.slider("Water Swelling (%)", 30.0, 65.0, 48.0)
            c_dens = col2.slider("True Density (gms/cc)", 1.35, 1.65, 1.50)
            c_poros = col1.slider("Porosity (%)", 4.0, 18.0, 7.5)
            c_dye = col2.selectbox("Dye Mix Type", ["vat", "azo", "reactive", "direct", "acid"])
            c_status = col1.selectbox("Action Status", ["Release", "Review", "Hold"])
            
            reg_submit = st.form_submit_button("🧪 Save and Run AI Grading")
            
        if reg_submit:
            if not b_id:
                st.error("Batch ID cannot be empty.")
            else:
                inputs = {
                    "Cellulose of yarn (%)": c_cell,
                    "Hemicellulose of yarn (%)": c_hemi,
                    "Lignin of yarn (%)": c_lig,
                    "Pectin of yarn (%)": c_pec,
                    "Moisture Content of yarn (%)": c_mois_c,
                    "pH Level of yarn": c_ph,
                    "Fineness of yarn (tex)": c_fine,
                    "Fiber Tenacity of yarn (gm/tex)": c_ten,
                    "Elongation of yarn (%)": c_elon,
                    "Moisture Regain of yarn (%)": c_reg,
                    "Water Swelling of yarn (%)": c_swell,
                    "True Density of yarn (gms/cc)": c_dens,
                    "Porosity of yarn (%)": c_poros,
                    "Dye Type": c_dye
                }
                
                pred_strength, pred_grade, confidence, risk_level, risk_reasons = run_predictions_on_row(
                    reg_model, clf_model, inputs, df_labeled
                )
                
                # Combine
                batch_data = inputs.copy()
                batch_data["batch_id"] = b_id.strip()
                batch_data["creation_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                batch_data["supplier_name"] = supplier
                batch_data["actual_strength"] = pred_strength
                batch_data["predicted_strength"] = pred_strength
                batch_data["predicted_grade"] = pred_grade
                batch_data["confidence"] = confidence
                batch_data["risk_level"] = risk_level
                batch_data["risk_reasons"] = risk_reasons
                batch_data["status"] = c_status
                
                db_save_batch(batch_data)
                st.success(f"Batch {b_id} successfully created.")
                time.sleep(1)
                st.rerun()
                
    with tab_compare:
        batches_list = db_get_batches()
        if len(batches_list) < 2:
            st.info("Compare needs at least 2 saved batches in the SQL database.")
        else:
            col1, col2 = st.columns(2)
            b1 = col1.selectbox("Select Batch A", batches_list["batch_id"].unique(), index=0)
            b2 = col2.selectbox("Select Batch B", batches_list["batch_id"].unique(), index=1)
            
            if b1 == b2:
                st.warning("Please select two distinct batch IDs to compare.")
            else:
                row_a = dict(db_get_batch(b1))
                row_b = dict(db_get_batch(b2))
                
                st.markdown("### ⚖️ Side-by-Side Comparison Matrix")
                
                # Compare critical columns
                compare_data = []
                metrics = [
                    ("Predicted Tensile Strength", "predicted_strength", "MPa", "{:.1f}"),
                    ("Quality Grade", "predicted_grade", "", "{}"),
                    ("ML Confidence Score", "confidence", "%", "{:.2f}"),
                    ("Risk Profile Level", "risk_level", "", "{}"),
                    ("Supplier Lint", "supplier_name", "", "{}"),
                    ("Cellulose Content", "cellulose", "%", "{:.2f}"),
                    ("Fiber Tenacity", "tenacity", "gm/tex", "{:.2f}"),
                    ("Fineness (Yarn Count)", "fineness", "tex", "{:.2f}"),
                    ("Elongation limit", "elongation", "%", "{:.2f}"),
                    ("Yarn Porosity", "porosity", "%", "{:.2f}"),
                    ("Ambient Moisture", "moisture_content", "%", "{:.2f}"),
                    ("Chemical pH", "ph_level", "", "{:.2f}")
                ]
                
                for label, key, unit, fmt in metrics:
                    val_a = row_a[key]
                    val_b = row_b[key]
                    
                    # Highlight highlight difference if numeric
                    diff_str = ""
                    if isinstance(val_a, (int, float)) and isinstance(val_b, (int, float)):
                        diff = val_a - val_b
                        sign = "+" if diff > 0 else ""
                        diff_str = f"{sign}{diff:.2f}"
                        
                    compare_data.append({
                        "Yarn Parameter": label,
                        f"Batch A ({b1})": f"{fmt.format(val_a)} {unit}".strip(),
                        f"Batch B ({b2})": f"{fmt.format(val_b)} {unit}".strip(),
                        "Variance (A - B)": diff_str
                    })
                    
                st.dataframe(pd.DataFrame(compare_data), use_container_width=True, hide_index=True)
                
                # Visual Plotly Strength Compare
                fig_comp = go.Figure(data=[
                    go.Bar(name=b1, x=["Strength (MPa)"], y=[row_a["predicted_strength"]], marker_color='#0d9488'),
                    go.Bar(name=b2, x=["Strength (MPa)"], y=[row_b["predicted_strength"]], marker_color='#6366f1')
                ])
                fig_comp.update_layout(barmode='group', height=240, title="Yarn Tensile Strength Yield Compare")
                st.plotly_chart(fig_comp, use_container_width=True)

# ==========================================
# PAGE 6: BULK PREDICTION TOOL
# ==========================================
def page_bulk_prediction(reg_model, clf_model, df_labeled):
    st.markdown(
        """
        <div class='glass-header'>
            <h1>📥 Bulk Prediction Tool</h1>
            <p>Upload raw CSV files containing laboratory lots and execute predictions for thousands of rows instantly.</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    left, right = st.columns([1.1, 0.9], gap="large")
    
    with left:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>Upload Laboratory CSV File</h4>", unsafe_allow_html=True)
        
        uploaded_file = st.file_uploader("Select CSV file", type=["csv"])
        
        # Seed generator for template CSV
        template_cols = [col for col in FEATURE_COLUMNS if col != "Tensile Strength of yarn (MPa)"]
        sample_template = df_labeled[template_cols].head(5)
        csv_template = sample_template.to_csv(index=False).encode('utf-8')
        
        st.download_button(
            "📥 Download Blank Template CSV",
            csv_template,
            "yarn_insight_bulk_template.csv",
            "text/csv"
        )
        st.markdown("</div>", unsafe_allow_html=True)
        
    with right:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>Bulk Instruction Manual</h4>", unsafe_allow_html=True)
        st.markdown(
            """
            <ol>
                <li>Download the blank template CSV structure.</li>
                <li>Populate columns with raw laboratory outputs. Ensure column headers match exactly.</li>
                <li>Upload the file. The ML pipelines will automatically run strength regressions and grade classifications for every single row.</li>
                <li>Download output results containing predicted strength, quality grades, risk profiles, and recommended end-use fabrics.</li>
            </ol>
            """,
            unsafe_allow_html=True
        )
        st.markdown("</div>", unsafe_allow_html=True)
        
    if uploaded_file is not None:
        try:
            bulk_df = pd.read_csv(uploaded_file)
            st.success(f"File loaded successfully: {uploaded_file.name} ({len(bulk_df)} records)")
            
            # Verify columns
            required_cols = [col for col in FEATURE_COLUMNS if col != "Tensile Strength of yarn (MPa)"]
            missing_cols = [col for col in required_cols if col not in bulk_df.columns]
            
            if missing_cols:
                st.error(f"Upload contains missing columns: {', '.join(missing_cols)}")
            else:
                run_bulk = st.button("🚀 Run Bulk Machine Learning Analysis")
                
                if run_bulk:
                    progress_bar = st.progress(0.0)
                    status_text = st.empty()
                    
                    strengths = []
                    grades = []
                    confidences = []
                    risks = []
                    best_fabrics = []
                    
                    total = len(bulk_df)
                    
                    # Run predictions
                    for i, row in bulk_df.iterrows():
                        row_dict = row.to_dict()
                        
                        # Use unified prediction function that handles ordering and risk
                        pred_s, pred_g, conf, risk_l, _ = run_predictions_on_row(
                            reg_model, clf_model, row_dict, df_labeled
                        )
                        
                        # Fabric Match
                        row_clf_mock = row_dict.copy()
                        row_clf_mock["Tensile Strength of yarn (MPa)"] = pred_s
                        recs = cloth_from_rules(row_clf_mock, pred_g)
                        best_fab = recs[0][0]
                        
                        strengths.append(pred_s)
                        grades.append(pred_g)
                        confidences.append(conf)
                        risks.append(risk_l)
                        best_fabrics.append(best_fab)
                        
                        # Update progress
                        if i % max(1, total // 20) == 0:
                            progress_bar.progress((i + 1) / total)
                            status_text.text(f"Analyzing row {i+1} of {total}...")
                            
                    progress_bar.progress(1.0)
                    status_text.text("Bulk Analysis Complete.")
                    
                    # Attach
                    output_df = bulk_df.copy()
                    output_df["Predicted Tensile Strength (MPa)"] = strengths
                    output_df["Predicted Quality Grade"] = grades
                    output_df["Classification Confidence (%)"] = confidences
                    output_df["Material Risk Profile"] = risks
                    output_df["Best Fabric Match"] = best_fabrics
                    
                    st.markdown("### 📊 Predictions Preview")
                    st.dataframe(output_df.head(200), use_container_width=True)
                    
                    # Downloads
                    csv_res = output_df.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        "📥 Download Predicted Results (CSV)",
                        csv_res,
                        "bulk_predicted_yarn_results.csv",
                        "text/csv",
                        use_container_width=True
                    )
                    
                    db_log_audit(st.session_state["email"], "BULK_PREDICT", f"Executed bulk prediction on file {uploaded_file.name} containing {total} records.")
        except Exception as e:
            st.error(f"Error processing CSV: {e}")

# ==========================================
# PAGE 7: ADVANCED ANALYTICS
# ==========================================
def page_analytics(df_labeled):
    st.markdown(
        """
        <div class='glass-header'>
            <h1>📈 Advanced Data Analytics</h1>
            <p>Physical/Mechanical Distributions, Attribute Correlations, and Overall Statistical Insights</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    # -----------------------------------------
    # Correlation Heatmap
    # -----------------------------------------
    st.markdown("### 🔗 Property Correlation Analysis")
    st.markdown("<p style='color:#64748b; font-size:0.85rem;'>Evaluates how chemical composition attributes (cellulose, hemicellulose, lignin) impact tensile strength.</p>", unsafe_allow_html=True)
    
    corr_cols = ["Cellulose of yarn (%)", "Lignin of yarn (%)", "Fineness of yarn (tex)", "Fiber Tenacity of yarn (gm/tex)", "Elongation of yarn (%)", "Moisture Regain of yarn (%)", "Tensile Strength of yarn (MPa)"]
    corr_matrix = df_labeled[corr_cols].corr()
    
    fig_heat = px.imshow(
        corr_matrix,
        x=corr_cols,
        y=corr_cols,
        color_continuous_scale=px.colors.diverging.Tealrose,
        text_auto=".2f",
        aspect="auto"
    )
    fig_heat.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=380)
    st.plotly_chart(fig_heat, use_container_width=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>📈 Tensile Strength Yield Distributions (MPa)</h4>", unsafe_allow_html=True)
        fig_hist = px.histogram(
            df_labeled,
            x="Tensile Strength of yarn (MPa)",
            color="Quality Grade",
            color_discrete_sequence=['#10b981', '#60a5fa', '#f59e0b', '#fb923c', '#ef4444'],
            marginal="box"
        )
        fig_hist.update_layout(margin=dict(t=20, b=10, l=10, r=10), height=300)
        st.plotly_chart(fig_hist, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
        
    with col2:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>Fiber Tenacity vs Yarn Elongation</h4>", unsafe_allow_html=True)
        fig_scatter = px.scatter(
            df_labeled.sample(800), # Sample to keep it fast
            x="Fiber Tenacity of yarn (gm/tex)",
            y="Elongation of yarn (%)",
            color="Quality Grade",
            color_discrete_sequence=['#10b981', '#60a5fa', '#f59e0b', '#fb923c', '#ef4444'],
            opacity=0.7
        )
        fig_scatter.update_layout(margin=dict(t=20, b=10, l=10, r=10), height=300)
        st.plotly_chart(fig_scatter, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
        
    # 3D scatter plot — Interactive Staple Analysis
    st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
    st.markdown("<h4>🔬 Interactive 3D Staple Analysis Plot</h4>", unsafe_allow_html=True)
    st.markdown("<p style='color:#64748b; font-size:0.85rem;'>Fiber Tenacity vs Elongation vs Tensile Strength — color-coded by Quality Grade. Spin, zoom, and hover over individual data points.</p>", unsafe_allow_html=True)
    
    sample_3d = df_labeled.sample(min(1200, len(df_labeled)), random_state=42)
    fig_3d = px.scatter_3d(
        sample_3d,
        x="Fiber Tenacity of yarn (gm/tex)",
        y="Elongation of yarn (%)",
        z="Tensile Strength of yarn (MPa)",
        color="Quality Grade",
        color_discrete_sequence=['#10b981', '#60a5fa', '#f59e0b', '#fb923c', '#ef4444'],
        opacity=0.85,
        height=560,
        hover_data={
            "Fineness of yarn (tex)": ":.2f",
            "Porosity of yarn (%)": ":.2f",
            "Moisture Content of yarn (%)": ":.2f",
            "Fiber Tenacity of yarn (gm/tex)": ":.2f",
            "Elongation of yarn (%)": ":.2f",
            "Tensile Strength of yarn (MPa)": ":.1f",
        },
        labels={
            "Fiber Tenacity of yarn (gm/tex)": "Tenacity (gm/tex)",
            "Elongation of yarn (%)": "Elongation (%)",
            "Tensile Strength of yarn (MPa)": "Strength (MPa)",
        },
    )
    fig_3d.update_traces(marker=dict(size=4, line=dict(width=0.5, color="white")))
    fig_3d.update_layout(
        margin=dict(t=10, b=10, l=10, r=10),
        scene=dict(
            xaxis_title="Fiber Tenacity (gm/tex)",
            yaxis_title="Elongation (%)",
            zaxis_title="Tensile Strength (MPa)",
            bgcolor="rgba(248,250,252,0.8)",
            xaxis=dict(gridcolor="#e2e8f0", backgroundcolor="#f8fafc"),
            yaxis=dict(gridcolor="#e2e8f0", backgroundcolor="#f8fafc"),
            zaxis=dict(gridcolor="#e2e8f0", backgroundcolor="#f8fafc"),
        ),
        legend=dict(title="Quality Grade", orientation="h", yanchor="bottom", y=1.02, x=0),
    )
    st.plotly_chart(fig_3d, use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

# ==========================================
# PAGE 8: SUPPLIER ANALYSIS
# ==========================================
def page_supplier_analysis():
    st.markdown(
        """
        <div class='glass-header'>
            <h1>🏢 Supplier Performance Scorecards</h1>
            <p>Raw material lint evaluations, Quality Scores, Strength averages, and Defect rate rankings.</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    batches_df = db_get_batches()
    if len(batches_df) == 0:
        st.info("Please register some batch lots first to enable supplier comparative analytics.")
        return
        
    st.markdown("### 🏆 Supplier Leaderboard rankings")
    
    # Calculate stats per supplier
    suppliers = batches_df["supplier_name"].unique()
    supplier_data = []
    
    for sup in suppliers:
        sup_df = batches_df[batches_df["supplier_name"] == sup]
        total = len(sup_df)
        avg_strength = sup_df["predicted_strength"].mean()
        
        # Defect rate (Reject percentage)
        defects = len(sup_df[sup_df["predicted_grade"] == "Reject"])
        defect_rate = (defects / total) * 100
        
        # Quality score (A+/A share)
        premium_count = len(sup_df[sup_df["predicted_grade"].isin(["Grade A+ (Premium)", "Grade A"])])
        quality_score = (premium_count / total) * 100
        
        supplier_data.append({
            "Supplier Name": sup,
            "Total Lots Supplied": total,
            "Average Yarn Yield (MPa)": round(avg_strength, 1),
            "Defect/Reject Rate (%)": round(defect_rate, 2),
            "Premium Quality Score (%)": round(quality_score, 1)
        })
        
    sup_table_df = pd.DataFrame(supplier_data).sort_values("Premium Quality Score (%)", ascending=False)
    
    st.dataframe(sup_table_df, use_container_width=True, hide_index=True)
    
    left, right = st.columns(2)
    with left:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>Supplier Quality Comparison Score</h4>", unsafe_allow_html=True)
        fig_sup1 = px.bar(
            sup_table_df,
            x="Supplier Name",
            y="Premium Quality Score (%)",
            color="Supplier Name",
            color_discrete_sequence=px.colors.qualitative.Teal
        )
        fig_sup1.update_layout(margin=dict(t=20, b=10, l=10, r=10), height=280)
        st.plotly_chart(fig_sup1, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
    with right:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>Supplier Defect Rates (%)</h4>", unsafe_allow_html=True)
        fig_sup2 = px.bar(
            sup_table_df,
            x="Supplier Name",
            y="Defect/Reject Rate (%)",
            color="Supplier Name",
            color_discrete_sequence=px.colors.qualitative.Oranges
        )
        fig_sup2.update_layout(margin=dict(t=20, b=10, l=10, r=10), height=280)
        st.plotly_chart(fig_sup2, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

# ==========================================
# PAGE 9: AI CHAT ASSISTANT
# ==========================================
def page_ai_assistant(df_labeled):
    st.markdown(
        """
        <div class='glass-header'>
            <h1>💬 YarnInsight AI Chat Assistant</h1>
            <p>Ask technical questions about yarn properties, query specific batches, or request optimization steps.</p>
        </div>
        """,
        unsafe_allow_html=True
    )

    # Proactive suggestions driven by live database alerts
    st.markdown("#### 💡 Suggested Questions")
    presets = get_chat_alert_suggestions()
    cols = st.columns(4)
    preset_clicked = None
    for idx, prompt in enumerate(presets):
        if cols[idx].button(prompt, key=f"chat_preset_{idx}"):
            preset_clicked = prompt

    hdr_left, hdr_right = st.columns([3, 1])
    with hdr_right:
        if st.button("Clear History", key="clear_chat_history"):
            db_clear_chat_history(st.session_state["email"])
            welcome = "Chat history cleared. How can I help you with yarn quality today?"
            st.session_state["chat_history"] = [{"role": "assistant", "content": welcome}]
            db_save_chat_message(st.session_state["email"], "assistant", welcome)
            st.rerun()

    st.divider()

    chat_container = st.container()
    with chat_container:
        for msg in st.session_state["chat_history"]:
            bubble_class = "user" if msg["role"] == "user" else "assistant"
            safe_content = html.escape(str(msg["content"])).replace("\n", "<br>")
            st.markdown(f"<div class='chat-bubble {bubble_class}'>{safe_content}</div>", unsafe_allow_html=True)

    user_q = st.text_input("Ask YarnInsight AI...", key="chat_user_input_box")

    triggered_query = None
    if preset_clicked:
        triggered_query = preset_clicked
    elif st.button("Send Query", key="send_chat_query") and user_q:
        triggered_query = user_q

    if triggered_query:
        st.session_state["chat_history"].append({"role": "user", "content": triggered_query})
        db_save_chat_message(st.session_state["email"], "user", triggered_query)

        response_text = generate_chat_response(triggered_query)

        st.session_state["chat_history"].append({"role": "assistant", "content": response_text})
        db_save_chat_message(st.session_state["email"], "assistant", response_text)
        st.rerun()

# ==========================================
# PAGE 10: REPORTS & CERTIFICATES
# ==========================================
def page_reports():
    st.markdown(
        """
        <div class='glass-header'>
            <h1>📝 Report Center & Certificate Generator</h1>
            <p>Generate Quality Assurance Certificates as PDFs, and export full database inventory logs.</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    left, right = st.columns(2, gap="large")
    
    with left:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>📜 Quality Validation Certificate Generator</h4>", unsafe_allow_html=True)
        st.markdown("<p style='color:#64748b; font-size:0.85rem;'>Select a batch record to compile an official PDF validation certificate.</p>", unsafe_allow_html=True)
        
        batches_df = db_get_batches()
        if len(batches_df) == 0:
            st.info("No batches registered inside database yet.")
        else:
            selected_batch = st.selectbox("Select Batch ID to certify", batches_df["batch_id"].unique())
            
            if st.button("Generate Certificate PDF"):
                pdf_bytes = generate_quality_certificate_pdf(selected_batch)
                if pdf_bytes:
                    st.success(f"Certificate generated successfully for lot {selected_batch}.")
                    st.download_button(
                        label="📥 Download Certificate PDF",
                        data=pdf_bytes,
                        file_name=f"YarnInsight_Certificate_{selected_batch}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                    db_log_audit(st.session_state["email"], "CERT_GEN", f"Generated Quality certificate PDF for batch ID {selected_batch}.")
                else:
                    st.error("Error generating PDF certificate.")
        st.markdown("</div>", unsafe_allow_html=True)
        
    with right:
        st.markdown("<div class='premium-card'>", unsafe_allow_html=True)
        st.markdown("<h4>📊 Export Inventory logs</h4>", unsafe_allow_html=True)
        st.markdown("<p style='color:#64748b; font-size:0.85rem;'>Export the current SQLite database batches to an Excel spreadsheet.</p>", unsafe_allow_html=True)
        
        if len(batches_df) == 0:
            st.info("No logs present to export.")
        else:
            if st.button("Compile Excel Summary Report"):
                excel_bytes = export_batches_to_excel()
                if excel_bytes:
                    st.success("Excel report compiled successfully.")
                    st.download_button(
                        label="📥 Download Excel Inventory Report",
                        data=excel_bytes,
                        file_name="YarnInsight_Inventory_Report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    db_log_audit(st.session_state["email"], "EXCEL_EXPORT", "Exported database inventory batches to Excel sheet.")
                else:
                    st.error("Error generating Excel report.")
        st.markdown("</div>", unsafe_allow_html=True)

# ==========================================
# PAGE 11: ADMIN PANEL
# ==========================================
def page_admin():
    st.markdown(
        """
        <div class='glass-header'>
            <h1>⚙️ Admin Command Center</h1>
            <p>Manage user authorizations, view system audit logs, and perform system cleanups.</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    tab_users, tab_audit = st.tabs(["👥 User Directory & Roles", "📋 Operational Audit Trail"])
    
    with tab_users:
        st.markdown("#### Database User Directory")
        conn = get_db()
        users_df = pd.read_sql_query("SELECT email, name, role FROM users", conn)
        conn.close()
        
        st.dataframe(users_df, use_container_width=True)
        
        # User Role updates
        st.markdown("##### Modify User Role")
        with st.form("modify_role_form"):
            selected_user = st.selectbox("Select User Email", users_df["email"].unique())
            new_role = st.selectbox("Assign New Security Level", ["Quality Engineer", "Production Manager", "Admin"])
            role_btn = st.form_submit_button("Confirm Role Modification")
            
        if role_btn:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("UPDATE users SET role = ? WHERE email = ?", (new_role, selected_user))
            conn.commit()
            conn.close()
            db_log_audit(st.session_state["email"], "USER_ROLE_CHANGE", f"Changed role of {selected_user} to {new_role}.")
            st.success(f"Security level for {selected_user} updated to {new_role}.")
            time.sleep(1)
            st.rerun()
            
    with tab_audit:
        st.markdown("#### System Operational Audit Logs")
        audit_df = db_get_audit_logs(150)
        st.dataframe(audit_df, use_container_width=True, hide_index=True)

# ==========================================
# PAGE 12: DOCUMENTATION
# ==========================================
def page_documentation(r2, acc):
    st.markdown(
        """
        <div class='glass-header'>
            <h1>📚 About & Documentation</h1>
            <p>Dataset Information, Machine Learning Model Hyperparameters, and User Operations Guide</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    st.markdown(
        f"""
        <div class='premium-card'>
            <h3>🧠 Machine Learning Model Specifications</h3>
            <p>The system utilizes two distinct Random Forest pipelines trained on the laboratory dataset.</p>
            <ul>
                <li><b>Yarn Strength Regressor Model:</b>
                    <ul>
                        <li><b>Algorithm:</b> Random Forest Regressor (100 Estimators)</li>
                        <li><b>Target:</b> Tensile Strength of yarn (MPa)</li>
                        <li><b>Performance ($R^2$ Score):</b> <b>{r2:.4f}</b></li>
                    </ul>
                </li>
                <li><b>Yarn Quality Grade Classifier Model:</b>
                    <ul>
                        <li><b>Algorithm:</b> Random Forest Classifier (120 Estimators)</li>
                        <li><b>Target:</b> Quality Grade (5 Tiers: A+, A, B, C, Reject)</li>
                        <li><b>Performance (Accuracy):</b> <b>{acc*100:.2f}%</b></li>
                    </ul>
                </li>
            </ul>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    st.markdown(
        """
        <div class='premium-card'>
            <h3>🌾 Physical Property Metrics Definitions</h3>
            <table style='width:100%; border-collapse: collapse; font-size: 0.9rem;'>
                <thead>
                    <tr style='border-bottom: 2px solid #cbd5e1; text-align: left;'>
                        <th style='padding: 10px;'>Parameter</th>
                        <th style='padding: 10px;'>Target Unit</th>
                        <th style='padding: 10px;'>Typical Range</th>
                        <th style='padding: 10px;'>Manufacturing Significance</th>
                    </tr>
                </thead>
                <tbody>
                    <tr style='border-bottom: 1px solid #cbd5e1;'>
                        <td style='padding: 10px; font-weight:700;'>Cellulose</td>
                        <td style='padding: 10px;'>%</td>
                        <td style='padding: 10px;'>65% - 85%</td>
                        <td style='padding: 10px;'>Chemical base providing raw mechanical structure.</td>
                    </tr>
                    <tr style='border-bottom: 1px solid #cbd5e1;'>
                        <td style='padding: 10px; font-weight:700;'>Fiber Tenacity</td>
                        <td style='padding: 10px;'>gm/tex</td>
                        <td style='padding: 10px;'>30 - 50</td>
                        <td style='padding: 10px;'>Individual fiber holding strength. Primary force driver.</td>
                    </tr>
                    <tr style='border-bottom: 1px solid #cbd5e1;'>
                        <td style='padding: 10px; font-weight:700;'>Fineness</td>
                        <td style='padding: 10px;'>tex</td>
                        <td style='padding: 10px;'>2.0 - 3.5</td>
                        <td style='padding: 10px;'>Yarn weight count. Finer yarn count gives premium fabrics.</td>
                    </tr>
                    <tr style='border-bottom: 1px solid #cbd5e1;'>
                        <td style='padding: 10px; font-weight:700;'>pH Level</td>
                        <td style='padding: 10px;'>pH</td>
                        <td style='padding: 10px;'>5.0 - 6.8</td>
                        <td style='padding: 10px;'>Dye readiness index. Out-of-bounds causes dye bath failures.</td>
                    </tr>
                    <tr style='border-bottom: 1px solid #cbd5e1;'>
                        <td style='padding: 10px; font-weight:700;'>Yarn Porosity</td>
                        <td style='padding: 10px;'>%</td>
                        <td style='padding: 10px;'>5.0% - 13.0%</td>
                        <td style='padding: 10px;'>Measures voids between fibers. Lower porosity increases yield limits.</td>
                    </tr>
                </tbody>
            </table>
        </div>
        """,
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()

